#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║         Tamil PSD Marketplace — GitHub Actions Pipeline v2          ║
║  ─────────────────────────────────────────────────────────────────  ║
║  1.  Auth Google Drive via OAuth2 refresh token (headless/CI)       ║
║  2.  Fetch designs.xlsx from Gurumoorthi repo                       ║
║  3.  Scan GDrive PSD folder — subfolders = categories               ║
║  4.  Skip already-processed IDs                                     ║
║  5.  Download PSD → convert to WebP preview (<80KB, watermarked)    ║
║  6.  Save WebP locally + upload to backend-sample/preview_image/    ║
║  7.  Zip PSD → upload to GDRIVE_UPLOAD_FOLDER → download URL        ║
║  8.  ModelScope Qwen2.5-VL vision AI → SEO title/tags/description   ║
║  9.  Append rows to designs.xlsx                                     ║
║  10. Push designs.xlsx back to Gurumoorthi repo                     ║
╚══════════════════════════════════════════════════════════════════════╝

Required GitHub Actions secrets:
  GOOGLE_CLIENT_ID      — Google OAuth2 client ID
  GOOGLE_CLIENT_SECRET  — Google OAuth2 client secret
  GOOGLE_REFRESH_TOKEN  — Google OAuth2 refresh token (one-time setup)
  GDRIVE_PSD_FOLDER     — Google Drive folder ID containing PSD subfolders
  GDRIVE_UPLOAD_FOLDER  — Google Drive folder ID for uploaded ZIP files
  SCOPE                 — ModelScope API token
  GH_PAT                — GitHub Personal Access Token (repo scope, both repos)
"""

import os, io, sys, re, time, json, base64, zipfile, traceback
import requests
import pandas as pd
from pathlib import Path
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ╔══════════════════════════════════════════════╗
# ║           CONFIG FROM SECRETS               ║
# ╚══════════════════════════════════════════════╝

GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")
GDRIVE_PSD_FOLDER    = os.environ.get("GDRIVE_PSD_FOLDER", "")
GDRIVE_UPLOAD_FOLDER = os.environ.get("GDRIVE_UPLOAD_FOLDER", "")
MODELSCOPE_TOKEN     = os.environ.get("SCOPE", "")
GH_PAT               = os.environ.get("GH_PAT", "")
GITHUB_TOKEN         = os.environ.get("GITHUB_TOKEN", "")  # built-in, for backend-sample

# ── Repositories ──────────────────────────────────────────────────────────────
BACKEND_REPO  = "moorthiguru33/backend-sample"    # preview images live here
CONTENT_REPO  = "moorthiguru33/Gurumoorthi"       # designs.xlsx lives here
PREVIEW_DIR   = "preview_image"
DESIGNS_PATH  = "designs.xlsx"

# Local directory for WebP files (committed via git in workflow step)
PREVIEW_LOCAL_DIR = Path(PREVIEW_DIR)

# jsDelivr CDN — images are served from backend-sample main branch
CDN_BASE = f"https://cdn.jsdelivr.net/gh/{BACKEND_REPO}@main/{PREVIEW_DIR}"

# ── ModelScope ────────────────────────────────────────────────────────────────
MODELSCOPE_API = "https://api-inference.modelscope.ai/v1/chat/completions"
VISION_MODEL   = "Qwen/Qwen2.5-VL-32B-Instruct"

# ── Processing limits ─────────────────────────────────────────────────────────
_env_count  = int(os.environ.get("FILE_COUNT", "0"))
MAX_PER_RUN = _env_count if _env_count > 0 else 50
AI_DELAY    = 5      # seconds between AI calls

# ── WebP preview settings ─────────────────────────────────────────────────────
PREVIEW_MAX_SIZE   = 1280
WEBP_TARGET_KB     = 80
WEBP_QUALITY_START = 82
WEBP_QUALITY_MIN   = 30

# ── Watermark ─────────────────────────────────────────────────────────────────
WATERMARK_TEXT    = "www.tamilpsd.in"
WATERMARK_OPACITY = 38

# ── Excel columns ─────────────────────────────────────────────────────────────
XLSX_HEADERS = [
    "ID", "Download URL", "Title", "Category", "Tags", "Description",
    "Dimensions", "DPI", "File Size", "Color Mode", "Software", "Fonts Used", "Preview URL"
]
COL_WIDTHS = [28, 70, 65, 20, 80, 110, 22, 10, 12, 14, 24, 32, 70]

FOLDER_MIME = "application/vnd.google-apps.folder"
IMAGE_EXTS  = {".psd", ".psb", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".webp"}


def log(msg: str):
    print(msg, flush=True)


# ╔══════════════════════════════════════════════╗
# ║         GOOGLE DRIVE AUTH (HEADLESS)        ║
# ╚══════════════════════════════════════════════╝

_token_cache = {"access_token": None, "expires_at": 0.0}


def gdrive_token() -> str:
    now = time.time()
    if _token_cache["access_token"] and now < _token_cache["expires_at"] - 60:
        return _token_cache["access_token"]
    resp = requests.post("https://oauth2.googleapis.com/token", data={
        "client_id":     GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "refresh_token": GOOGLE_REFRESH_TOKEN,
        "grant_type":    "refresh_token",
    }, timeout=30)
    data = resp.json()
    if "access_token" not in data:
        raise RuntimeError(f"GDrive OAuth2 failed: {data}")
    _token_cache["access_token"] = data["access_token"]
    _token_cache["expires_at"]   = now + data.get("expires_in", 3600)
    log("  🔑 GDrive token refreshed")
    return _token_cache["access_token"]


def gdrive_headers() -> dict:
    return {"Authorization": f"Bearer {gdrive_token()}"}


# ╔══════════════════════════════════════════════╗
# ║         GOOGLE DRIVE OPERATIONS             ║
# ╚══════════════════════════════════════════════╝

def gdrive_list_folder(folder_id: str) -> list:
    items, page_token = [], None
    while True:
        params = {
            "q":        f"'{folder_id}' in parents and trashed=false",
            "fields":   "nextPageToken,files(id,name,mimeType,size)",
            "pageSize": "1000",
        }
        if page_token:
            params["pageToken"] = page_token
        resp = requests.get(
            "https://www.googleapis.com/drive/v3/files",
            params=params, headers=gdrive_headers(), timeout=30
        )
        data = resp.json()
        items.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    return items


def gdrive_download(file_id: str) -> bytes:
    resp = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={"alt": "media"},
        headers=gdrive_headers(),
        timeout=300,
        stream=True
    )
    resp.raise_for_status()
    return resp.content


def gdrive_file_exists(filename: str, folder_id: str):
    safe = filename.replace("'", "\\'")
    resp = requests.get(
        "https://www.googleapis.com/drive/v3/files",
        params={
            "q":        f"name='{safe}' and '{folder_id}' in parents and trashed=false",
            "fields":   "files(id,name)",
            "pageSize": "1",
        },
        headers=gdrive_headers(), timeout=20
    )
    files = resp.json().get("files", [])
    return files[0]["id"] if files else None


def gdrive_upload(file_bytes: bytes, filename: str, folder_id: str,
                  mime: str = "application/zip") -> str:
    existing = gdrive_file_exists(filename, folder_id)
    if existing:
        log(f"    ⏭  GDrive already exists: {filename}")
        return f"https://drive.usercontent.google.com/download?id={existing}&export=download&confirm=t"

    boundary = "==gdrive_boundary=="
    meta_json = json.dumps({"name": filename, "parents": [folder_id]}).encode("utf-8")
    body = (
        f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n".encode()
        + meta_json
        + f"\r\n--{boundary}\r\nContent-Type: {mime}\r\n\r\n".encode()
        + file_bytes
        + f"\r\n--{boundary}--".encode()
    )
    resp = requests.post(
        "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart",
        data=body,
        headers={
            **gdrive_headers(),
            "Content-Type": f"multipart/related; boundary={boundary}",
        },
        timeout=600
    )
    data = resp.json()
    file_id = data.get("id", "")
    if not file_id:
        log(f"    ❌ GDrive upload error: {data}")
        return ""

    requests.post(
        f"https://www.googleapis.com/drive/v3/files/{file_id}/permissions",
        json={"type": "anyone", "role": "reader"},
        headers=gdrive_headers(), timeout=20
    )
    url = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
    log(f"    ✅ GDrive upload OK: {filename} → {file_id}")
    return url


def gdrive_find_or_create_folder(folder_name: str, parent_folder_id: str) -> str:
    safe = folder_name.replace("'", "\\'")
    resp = requests.get(
        "https://www.googleapis.com/drive/v3/files",
        params={
            "q": (
                f"name='{safe}' and '{parent_folder_id}' in parents "
                f"and mimeType='{FOLDER_MIME}' and trashed=false"
            ),
            "fields":   "files(id,name)",
            "pageSize": "1",
        },
        headers=gdrive_headers(), timeout=20
    )
    files = resp.json().get("files", [])
    if files:
        return files[0]["id"]

    resp = requests.post(
        "https://www.googleapis.com/drive/v3/files",
        json={
            "name":     folder_name,
            "mimeType": FOLDER_MIME,
            "parents":  [parent_folder_id],
        },
        headers=gdrive_headers(), timeout=20
    )
    data = resp.json()
    folder_id = data.get("id", "")
    if folder_id:
        log(f"    📁 Created 'final' folder inside parent {parent_folder_id[:12]}…")
    else:
        log(f"    ❌ Failed to create folder '{folder_name}': {data}")
    return folder_id


def gdrive_move_to_final(file_id: str, category_folder_id: str) -> bool:
    final_folder_id = gdrive_find_or_create_folder("final", category_folder_id)
    if not final_folder_id:
        return False

    resp = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={"fields": "parents"},
        headers=gdrive_headers(), timeout=20
    )
    if resp.status_code != 200:
        log(f"    ❌ gdrive_move_to_final: could not fetch parents (HTTP {resp.status_code})")
        return False
    current_parents = ",".join(resp.json().get("parents", []))

    resp = requests.patch(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={
            "addParents":    final_folder_id,
            "removeParents": current_parents,
            "fields":        "id,parents",
        },
        headers={**gdrive_headers(), "Content-Type": "application/json"},
        json={},
        timeout=30
    )
    if resp.status_code == 200:
        return True
    log(f"    ❌ gdrive_move_to_final failed (HTTP {resp.status_code}): {resp.text[:200]}")
    return False


def scan_gdrive_structure(root_id: str) -> list:
    result = []
    items = gdrive_list_folder(root_id)
    for item in items:
        name = item["name"]
        fid  = item["id"]
        mime = item["mimeType"]
        if mime == FOLDER_MIME:
            if name.lower() == "final":
                continue
            sub_items = gdrive_list_folder(fid)
            for sub in sub_items:
                if sub["mimeType"] == FOLDER_MIME:
                    continue
                ext = Path(sub["name"]).suffix.lower()
                if ext in IMAGE_EXTS:
                    result.append((sub["id"], sub["name"], name, fid))
        else:
            ext = Path(name).suffix.lower()
            if ext in IMAGE_EXTS:
                result.append((fid, name, "", root_id))
    return result


# ╔══════════════════════════════════════════════╗
# ║         WATERMARK                           ║
# ╚══════════════════════════════════════════════╝

def add_watermark(img: PILImage.Image) -> PILImage.Image:
    if img.mode != "RGBA":
        img = img.convert("RGBA")

    w, h = img.size
    overlay = PILImage.new("RGBA", (w, h), (0, 0, 0, 0))
    draw    = ImageDraw.Draw(overlay)

    font_size = max(14, int(min(w, h) * 0.025))
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
    except Exception:
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", font_size)
        except Exception:
            font = ImageFont.load_default()

    bbox  = draw.textbbox((0, 0), WATERMARK_TEXT, font=font)
    txt_w = bbox[2] - bbox[0]
    txt_h = bbox[3] - bbox[1]

    x_step = txt_w + max(40, int(w * 0.12))
    y_step = txt_h + max(30, int(h * 0.10))

    for y_start in range(-h, h * 2, y_step):
        for x_start in range(-w, w * 2, x_step):
            draw.text(
                (x_start, y_start),
                WATERMARK_TEXT,
                font=font,
                fill=(255, 255, 255, WATERMARK_OPACITY)
            )

    overlay_rot = overlay.rotate(-30, expand=False, resample=PILImage.BICUBIC)
    combined    = PILImage.alpha_composite(img, overlay_rot)
    return combined


# ╔══════════════════════════════════════════════╗
# ║         IMAGE CONVERSION → WebP             ║
# ╚══════════════════════════════════════════════╝

def to_webp_bytes(file_bytes: bytes, filename: str,
                  max_size: int = PREVIEW_MAX_SIZE) -> bytes:
    ext = Path(filename).suffix.lower()
    try:
        if ext in (".psd", ".psb"):
            try:
                from psd_tools import PSDImage
                psd = PSDImage.open(io.BytesIO(file_bytes))
                img = psd.composite()
                if img is None:
                    raise ValueError("psd_tools composite returned None")
            except Exception as psd_err:
                log(f"    ⚠  psd_tools: {psd_err} — falling back to PIL")
                img = PILImage.open(io.BytesIO(file_bytes))
        else:
            img = PILImage.open(io.BytesIO(file_bytes))

        if img.mode not in ("RGB", "RGBA", "L"):
            img = img.convert("RGB")

        if max(img.size) > max_size:
            img = img.copy()
            img.thumbnail((max_size, max_size), PILImage.LANCZOS)

        if img.mode == "L":
            img = img.convert("RGB")

        img = add_watermark(img)

        bg = PILImage.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[3])
        img = bg

        target_bytes = WEBP_TARGET_KB * 1024
        quality      = WEBP_QUALITY_START

        while quality >= WEBP_QUALITY_MIN:
            buf = io.BytesIO()
            img.save(buf, "WEBP", quality=quality, method=4)
            webp_data = buf.getvalue()
            if len(webp_data) <= target_bytes:
                break
            quality -= 5

        log(f"    ✅ WebP: {format_size(len(webp_data))} @ quality={quality}")
        return webp_data

    except Exception as e:
        log(f"    ❌ WebP conversion failed for {filename}: {e}")
        traceback.print_exc()
        return b""


def format_size(n_bytes: int) -> str:
    if n_bytes < 1024:
        return f"{n_bytes} B"
    elif n_bytes < 1_048_576:
        return f"{n_bytes/1024:.1f} KB"
    return f"{n_bytes/1_048_576:.2f} MB"


# ╔══════════════════════════════════════════════╗
# ║         GITHUB OPERATIONS                   ║
# ╚══════════════════════════════════════════════╝

def _gh_headers(token: str) -> dict:
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}


def github_file_sha(repo: str, path: str, token: str):
    resp = requests.get(
        f"https://api.github.com/repos/{repo}/contents/{path}",
        headers=_gh_headers(token), timeout=30
    )
    return resp.json().get("sha") if resp.status_code == 200 else None


def github_upload_file(repo: str, path: str, content: bytes,
                       token: str, commit_msg: str) -> bool:
    """
    Upload (or update) a file on GitHub via API.
    Returns True if file already exists OR upload succeeds.
    Does NOT skip if file already exists — checks for real content match.
    """
    sha = github_file_sha(repo, path, token)
    if sha:
        log(f"    ⏭  GitHub already exists (API): {path}")
        return True

    payload = {
        "message": commit_msg,
        "content": base64.b64encode(content).decode(),
    }
    resp = requests.put(
        f"https://api.github.com/repos/{repo}/contents/{path}",
        headers=_gh_headers(token), json=payload, timeout=60
    )
    if resp.status_code in (200, 201):
        log(f"    ✅ GitHub API upload OK: {path}")
        return True
    log(f"    ⚠  GitHub API upload failed {resp.status_code}: {resp.text[:200]}")
    return False


def save_webp_locally(webp_bytes: bytes, filename: str) -> bool:
    """
    Save WebP to the local preview_image/ directory.
    This file will be committed to the repo by the workflow git step.
    Returns True on success.
    """
    try:
        PREVIEW_LOCAL_DIR.mkdir(parents=True, exist_ok=True)
        out_path = PREVIEW_LOCAL_DIR / filename
        if out_path.exists():
            log(f"    ⏭  Local preview already exists: {filename}")
            return True
        out_path.write_bytes(webp_bytes)
        log(f"    ✅ Saved locally: {out_path} ({format_size(len(webp_bytes))})")
        return True
    except Exception as e:
        log(f"    ❌ Local save failed for {filename}: {e}")
        return False


def fetch_designs_xlsx(token: str):
    url  = f"https://api.github.com/repos/{CONTENT_REPO}/contents/{DESIGNS_PATH}"
    resp = requests.get(url, headers=_gh_headers(token), timeout=30)

    if resp.status_code == 404:
        log("  ℹ️  designs.xlsx not found in repo — will create fresh")
        return pd.DataFrame(columns=XLSX_HEADERS), None

    if resp.status_code != 200:
        log(f"  ⚠️  designs.xlsx fetch returned HTTP {resp.status_code} — starting fresh")
        return pd.DataFrame(columns=XLSX_HEADERS), None

    data = resp.json()
    sha  = data.get("sha")

    raw_b64 = data.get("content", "")
    raw = None

    if not raw_b64:
        download_url_direct = data.get("download_url", "")
        if download_url_direct:
            log("  📥 designs.xlsx is large (>1MB) — fetching via direct download URL...")
            try:
                dl_resp = requests.get(
                    download_url_direct, headers=_gh_headers(token), timeout=120
                )
                if dl_resp.status_code == 200:
                    raw = dl_resp.content
                    log(f"  ✅ Direct download OK: {len(raw)} bytes")
                else:
                    log(f"  ⚠️  Direct download failed: HTTP {dl_resp.status_code} — starting fresh")
                    return pd.DataFrame(columns=XLSX_HEADERS), None
            except Exception as dl_err:
                log(f"  ⚠️  Direct download error: {dl_err} — starting fresh")
                return pd.DataFrame(columns=XLSX_HEADERS), None
        else:
            log("  ⚠️  designs.xlsx API response had no content field — starting fresh")
            return pd.DataFrame(columns=XLSX_HEADERS), None

    if raw is None:
        try:
            raw = base64.b64decode(raw_b64.replace("\n", ""))
        except Exception as b64_err:
            log(f"  ⚠️  designs.xlsx base64 decode failed: {b64_err} — starting fresh")
            return pd.DataFrame(columns=XLSX_HEADERS), None

    if not raw[:4] == b"PK\x03\x04":
        log(f"  ⚠️  designs.xlsx bytes do not start with ZIP magic (got {raw[:4]!r}) — starting fresh")
        return pd.DataFrame(columns=XLSX_HEADERS), None

    try:
        df = pd.read_excel(io.BytesIO(raw), engine="openpyxl", dtype=str).fillna("")
    except Exception as xl_err:
        log(f"  ⚠️  designs.xlsx could not be parsed: {xl_err} — starting fresh")
        return pd.DataFrame(columns=XLSX_HEADERS), None

    for col in XLSX_HEADERS:
        if col not in df.columns:
            df[col] = ""
    df = df[XLSX_HEADERS]
    for col in XLSX_HEADERS:
        df[col] = df[col].astype(str).replace("nan", "")
    log(f"  ✅ designs.xlsx: {len(df)} rows, SHA {str(sha)[:10]}…")
    return df, sha


def push_designs_xlsx(df: pd.DataFrame, sha, token: str, commit_msg: str):
    url = f"https://api.github.com/repos/{CONTENT_REPO}/contents/{DESIGNS_PATH}"
    payload = {
        "message": commit_msg,
        "content": base64.b64encode(df_to_xlsx_bytes(df)).decode(),
    }
    if sha:
        payload["sha"] = sha
    resp = requests.put(url, headers=_gh_headers(token), json=payload, timeout=60)
    if resp.status_code in (200, 201):
        new_sha = resp.json().get("content", {}).get("sha")
        log(f"  ✅ designs.xlsx pushed! SHA {str(new_sha)[:10]}…")
        return new_sha
    log(f"  ❌ Push failed {resp.status_code}: {resp.text[:300]}")
    return sha


# ╔══════════════════════════════════════════════╗
# ║         EXCEL HELPER                        ║
# ╚══════════════════════════════════════════════╝

HEADER_FILL = PatternFill("solid", fgColor="0D1117")
HEADER_FONT = Font(color="00D4FF", bold=True, name="Courier New", size=11)
ROW_FILL_A  = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_B  = PatternFill("solid", fgColor="EBF3FB")
URL_FONT    = Font(color="0563C1", underline="single")
BORDER      = Border(bottom=Side(style="thin", color="D0D0D0"),
                     right= Side(style="thin", color="D0D0D0"))


def df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PSD Data"
    for ci, (hdr, w) in enumerate(zip(XLSX_HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 24
    for ri, (_, row_data) in enumerate(df.iterrows(), 2):
        fill = ROW_FILL_B if ri % 2 == 0 else ROW_FILL_A
        for ci, col in enumerate(XLSX_HEADERS, 1):
            val = row_data.get(col, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci in (2, 13) and val and str(val).startswith("http"):
                cell.font = URL_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(XLSX_HEADERS))}1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ╔══════════════════════════════════════════════╗
# ║         AI SEO PROMPT                       ║
# ╚══════════════════════════════════════════════╝

# FIX: Reduced description target from 400-500 words to 300-380 words.
# max_tokens reduced from 2000 → 900 to prevent runaway generation.
SEO_PROMPT = """Tamil PSD marketplace SEO expert.

⚠️ MANDATORY OUTPUT FORMAT — YOU MUST USE THESE EXACT MARKERS OR OUTPUT IS INVALID:
##TITLE##
##TAGS##
##DIMENSIONS##
##DESCRIPTION##

Begin your response DIRECTLY with ##TITLE## — no preamble, no explanation, no numbering.

CLASSIFY (silent): A=Person/cutout B=Text-only C=Single object | else=standard PSD
A: Title "[Name] PNG Cutout Download" | no printed name→use FOLDER name from 👤 OVERRIDE
B: Transliterate Tamil→English. Title "[Text] Title PNG Download"
C: Title "[Object] PNG Download"

RULES:
- Read ALL text first — confirms occasion/party/person. Never guess from color/face alone.
- Visiting card≠banner. Death≠festival. Be precise.
- ⛔ INVITATION BAN: NEVER use "invitation/invite/invitations/amandhippu/nimantrana" ANYWHERE unless "💌 FOLDER CONTEXT OVERRIDE — INVITATION FOLDER" appears above. Banners are BANNERS, not invitations.

##TITLE##
12–18 words, Title Case. End: "PSD Template Download" / "PNG Download" / "PSD Download".
Include: design type + occasion + key visual element + style/color + target audience or region.
Example: "Elegant Tamil Hindu Wedding Gold Border Flex Banner PSD Template Download"
UNIQUE per design — vary wording, visual element, and structure every time.

##TAGS##
10 unique lowercase comma-separated tags. Include: occasion, format, "psd template", 2–3 Tamil transliterated terms. No "free download". No numbering. No brackets.

##DIMENSIONS##
ONE only: 1800x1200 pixels | 1050x1500 pixels | 1080x1080 pixels | 2480x3508 pixels | 1050x600 pixels

##DESCRIPTION##
STRICT: 300–380 WORDS TOTAL. Count carefully — stop when you reach 380 words.
UNIQUE per design. Describe ONLY what is visible — concept, style, colors, usage.
No Photoshop/file/DPI/software/password mentions. No generic filler. No bullet lists.

Para 1 (70–80w): Design concept and mood — theme, cultural context, overall feel.
Para 2 (90–100w): Exact colors by name, background style, border/frame details, decorative motifs, patterns, and typography style — based on what is actually visible.
Para 3 (70–80w): Layout details — element arrangement, photo placeholder position, text zones, decorative accents, and spacing.
Para 4 (70–80w): Who uses this, for what specific purpose, and what occasions it fits perfectly.

REMINDER: Stop at 380 words. Your entire response MUST start with ##TITLE##.
"""


# ╔══════════════════════════════════════════════╗
# ║         FOLDER CONTEXT DETECTION            ║
# ╚══════════════════════════════════════════════╝

KNOWN_POLITICIANS = [
    "stalin", "mk stalin", "udhayanidhi", "kalaignar", "karunanidhi",
    "edappadi", "eps", "jayalalitha", "amma", "vijay", "thalapathy vijay",
    "tvk vijay", "modi", "narendra modi", "rahul gandhi", "annamalai",
    "k annamalai", "seeman", "thirumavalavan", "ramadoss", "anbumani",
    "cm", "minister", "mla", "mp",
]

KNOWN_PARTIES = {
    "dmk":      ("DMK", "DMK (Dravida Munnetra Kazhagam)", "rising sun symbol, red+black"),
    "admk":     ("ADMK", "ADMK (All India Anna Dravida Munnetra Kazhagam)", "two leaves, green+white"),
    "aiadmk":   ("ADMK", "ADMK (All India Anna Dravida Munnetra Kazhagam)", "two leaves, green+white"),
    "tvk":      ("TVK", "TVK (Tamilaga Vettri Kazhagam)", "TVK logo, blue, led by Vijay"),
    "bjp":      ("BJP", "BJP (Bharatiya Janata Party)", "lotus symbol, saffron+white"),
    "congress": ("Congress", "INC (Indian National Congress)", "open hand, tricolor"),
    "inc":      ("Congress", "INC (Indian National Congress)", "open hand, tricolor"),
    "vck":      ("VCK", "VCK (Viduthalai Chiruthaigal Katchi)", "fist symbol, blue+red"),
    "pmk":      ("PMK", "PMK (Pattali Makkal Katchi)", "mango, yellow+green"),
    "ntk":      ("NTK", "NTK (Nam Tamilar Katchi)", "tiger symbol, red"),
    "mmk":      ("MMK", "MMK (Manithaneya Makkal Katchi)", "yellow+red"),
    "mnm":      ("Political", "MNM (Makkal Needhi Maiam)", "scales symbol"),
}

FOLDER_TYPE_KEYWORDS = {
    "elements": "ELEMENTS", "element": "ELEMENTS", "png": "ELEMENTS",
    "clipart": "ELEMENTS", "cliparts": "ELEMENTS", "stickers": "ELEMENTS",
    "pamphlet": "PAMPHLET", "pamphlets": "PAMPHLET", "brochure": "PAMPHLET",
    "brochures": "PAMPHLET", "leaflet": "PAMPHLET", "flyer": "PAMPHLET",
    "death": "DEATH", "memorial": "DEATH", "condolence": "DEATH",
    "rip": "DEATH", "anjali": "DEATH",
    "wedding": "WEDDING", "marriage": "WEDDING", "thirumanam": "WEDDING",
    "kalyanam": "WEDDING",
    "birthday": "BIRTHDAY", "bday": "BIRTHDAY",
    "festival": "FESTIVAL", "pongal": "FESTIVAL", "diwali": "FESTIVAL",
    "business": "BUSINESS", "shop": "BUSINESS", "company": "BUSINESS",
    "visiting": "VISITING_CARD", "visitingcard": "VISITING_CARD",
    "invitation": "INVITATION", "invitations": "INVITATION",
    "invite": "INVITATION", "nimantrana": "INVITATION",
}

_FOLDER_HINTS = {
    "ELEMENTS": {
        "ai_hint": "🖼️ FOLDER CONTEXT OVERRIDE — ELEMENTS/PNG FOLDER.\nThis is a PNG element/clipart. Title: '[Object] PNG Element Free Download'. Tags MUST include 'png element', 'transparent png', 'clipart'. End title with 'PNG Free Download' NOT 'PSD Template Download'.",
        "category_hint": "Object PNG",
    },
    "PAMPHLET": {
        "ai_hint": "📄 FOLDER CONTEXT OVERRIDE — PAMPHLET/BROCHURE FOLDER.\nTitle format: '[Business/Topic] Pamphlet PSD Template Download'. Tags: 'pamphlet psd', 'brochure template', 'leaflet design'.",
        "category_hint": "Brochure/Pamphlet",
    },
    "DEATH": {
        "ai_hint": "🕊️ FOLDER CONTEXT OVERRIDE — DEATH/MEMORIAL FOLDER.\nThis is a death announcement or condolence banner. Use respectful, solemn language. NEVER call it an invitation.",
        "category_hint": "Condolence",
    },
    "WEDDING": {
        "ai_hint": "💒 FOLDER CONTEXT OVERRIDE — WEDDING FOLDER.\nThis is a wedding design. ⛔ Do NOT use 'invitation' — this is a banner/template.",
        "category_hint": "Wedding",
    },
    "BIRTHDAY": {
        "ai_hint": "🎂 FOLDER CONTEXT OVERRIDE — BIRTHDAY FOLDER.\nThis is a birthday design. ⛔ Do NOT use 'invitation' — this is a banner/template.",
        "category_hint": "Birthday",
    },
    "INVITATION": {
        "ai_hint": "💌 FOLDER CONTEXT OVERRIDE — INVITATION FOLDER.\nThis IS an invitation design. You MAY use the word 'invitation' in title, tags, and description.",
        "category_hint": "Invitation",
    },
    "VISITING_CARD": {
        "ai_hint": "💳 FOLDER CONTEXT OVERRIDE — VISITING CARD FOLDER.\nTitle: '[Profession] Visiting Card PSD Template Download'. Tags: 'visiting card psd', 'business card template'.",
        "category_hint": "Visiting Card",
    },
    "FESTIVAL": {
        "ai_hint": "🎉 FOLDER CONTEXT OVERRIDE — FESTIVAL FOLDER.\nThis is a festival design. Include the festival name in title and tags.",
        "category_hint": "Festival",
    },
    "BUSINESS": {
        "ai_hint": "🏪 FOLDER CONTEXT OVERRIDE — BUSINESS/SHOP FOLDER.\nThis is a business design. Title: '[Business Type] [Design Type] PSD Template Download'.",
        "category_hint": "Shop/Business",
    },
}


def get_folder_context(subfolder: str) -> dict:
    if not subfolder:
        return {"folder_type": "GENERIC", "ai_hint": "", "category_hint": ""}

    sf_lower   = subfolder.lower().strip().replace("-", " ").replace("_", " ")
    sf_nospace = sf_lower.replace(" ", "")

    for key, party in KNOWN_PARTIES.items():
        if sf_nospace == key.replace(" ", "") or bool(re.search(
            r'(?<![a-z])' + re.escape(key) + r'(?![a-z])', sf_lower
        )):
            return {
                "folder_type": f"PARTY_{key.upper()}",
                "ai_hint": (
                    f"⭐ FOLDER CONTEXT OVERRIDE — THIS IS A **{party[1].upper()}** FOLDER ⭐\n\n"
                    f"All images in this folder belong to {party[1]}.\n"
                    f"Party visual identity: {party[2]}\n"
                    f"Title MUST include party name. Tags MUST include party name and leader.\n"
                    f"Description: for {party[1]} political workers, election campaigns, flex banners.\n"
                    f"⚠️ Do NOT use the word 'invitation' — party designs are banners/posters."
                ),
                "category_hint": party[0],
            }

    sf_words = set(sf_lower.split())
    for pol in KNOWN_POLITICIANS:
        if set(pol.split()) & sf_words or pol in sf_lower:
            return {
                "folder_type": "PERSON",
                "ai_hint": (
                    f"👤 FOLDER CONTEXT OVERRIDE — PERSON FOLDER: \"{subfolder}\"\n\n"
                    f"All images are PNG cutouts/art of: {subfolder}\n"
                    f"Title MUST include name \"{subfolder}\".\n"
                    f"Write: \"{subfolder} PNG Cutout Download\" or \"{subfolder} HD PNG Download\".\n"
                    f"Tags: person's name, 'png cutout', 'transparent background', 'hd png'.\n"
                    f"⚠️ NEVER use 'Unknown Person' — always use: \"{subfolder}\"."
                ),
                "category_hint": "Person PNG",
            }

    for keyword, ftype in FOLDER_TYPE_KEYWORDS.items():
        if keyword in sf_lower:
            hint_data = _FOLDER_HINTS.get(ftype, {"ai_hint": "", "category_hint": ""})
            return {
                "folder_type": ftype,
                "ai_hint":     hint_data["ai_hint"],
                "category_hint": hint_data["category_hint"],
            }

    if subfolder and subfolder[0].isupper() and len(sf_lower.split()) <= 4:
        return {
            "folder_type": "PERSON",
            "ai_hint": (
                f"👤 FOLDER CONTEXT OVERRIDE — PERSON FOLDER: \"{subfolder}\"\n\n"
                f"Title MUST include \"{subfolder}\". "
                f"Write: \"{subfolder} PNG Cutout Download\".\n"
                f"NEVER write 'Unknown Person'."
            ),
            "category_hint": "Person PNG",
        }

    return {"folder_type": "GENERIC", "ai_hint": "", "category_hint": ""}


# ╔══════════════════════════════════════════════╗
# ║         RESPONSE PARSING + CLEANING         ║
# ╚══════════════════════════════════════════════╝

def clean_title(title: str) -> str:
    """
    Clean and validate SEO title.
    - Remove markdown bold/italic markers
    - Remove label prefixes
    - Remove surrounding quotes
    - Ensure ends with a valid download suffix
    """
    if not title:
        return title

    # Remove markdown bold/italic
    title = re.sub(r'\*+', '', title).strip()
    # Remove surrounding quotes
    title = title.strip("\"'")
    # Remove numbering at start (e.g. "1. Title" or "1) Title")
    title = re.sub(r'^\d+[\.\)]\s*', '', title).strip()
    # Remove label prefixes
    for pfx in ["Title:", "SEO Title:", "Output:", "Answer:", "TITLE:", "##TITLE##"]:
        if title.lower().startswith(pfx.lower()):
            title = title[len(pfx):].strip()
    # Remove surrounding brackets
    for ch in ["[", "]", "(", ")"]:
        title = title.replace(ch, "")
    title = title.strip()

    # Ensure ends with a recognised download suffix
    valid_endings = [
        "PSD Template Download", "PNG Download", "PSD Download",
        "Template Download", "Free Download", "Download"
    ]
    if not any(title.lower().endswith(e.lower()) for e in valid_endings):
        title = title.rstrip(".").strip() + " PSD Template Download"

    return title


def clean_tags(tags: str) -> str:
    """
    Clean SEO tags:
    - Split by comma, strip each tag
    - Lowercase
    - Remove duplicates
    - Remove numbering/markdown artifacts
    - Limit to 12 tags
    """
    if not tags:
        return tags

    # Remove markdown bold/italic
    tags = re.sub(r'\*+', '', tags)
    # Remove numbering (e.g. "1. tag, 2. tag")
    tags = re.sub(r'\b\d+[\.\)]\s*', '', tags)

    tag_list = [t.strip().lower().strip("\"'[]()") for t in tags.split(",")]
    tag_list = [t for t in tag_list if t and len(t) > 1]

    # Deduplicate preserving order
    seen, clean = set(), []
    for t in tag_list:
        if t not in seen:
            seen.add(t)
            clean.append(t)

    return ", ".join(clean[:12])


def parse_response(raw: str):
    """Parse structured AI response into (title, tags, dims, desc)."""
    def between(text, start, end=None):
        s = text.find(start)
        if s == -1:
            return ""
        s += len(start)
        if end:
            e = text.find(end, s)
            return text[s:e].strip() if e != -1 else text[s:].strip()
        return text[s:].strip()

    title = between(raw, "##TITLE##",       "##TAGS##")
    tags  = between(raw, "##TAGS##",        "##DIMENSIONS##")
    dims  = between(raw, "##DIMENSIONS##",  "##DESCRIPTION##")
    desc  = between(raw, "##DESCRIPTION##")

    # Fallback: AI sometimes writes **TITLE** or *TITLE* instead of ##TITLE##
    if not title:
        for pat in [r"\*{1,2}TITLE\*{1,2}[:\s]+(.+)", r"Title[:\s]+(.+)", r"^([A-Z].{10,80}(?:Download|Template|PSD))"]:
            m = re.search(pat, raw, re.IGNORECASE | re.MULTILINE)
            if m:
                title = m.group(1).strip()
                break

    # Fallback: first non-empty substantial line
    if not title:
        for line in raw.split("\n"):
            line = line.strip().strip("#").strip("*").strip()
            if len(line) > 15:
                title = line
                break

    # Fallback: tags from comma-separated line
    if not tags:
        for line in raw.split("\n"):
            if "," in line and len(line.split(",")) >= 4:
                stripped = line.strip().strip("#").strip("*").strip()
                if not stripped.startswith("Para") and len(stripped) < 300:
                    tags = stripped
                    break

    # Apply cleaning
    title = clean_title(title)
    tags  = clean_tags(tags)

    dims_m = re.search(r"\d+\s*[xX×]\s*\d+\s*pixels?", dims)
    if dims_m:
        dims = dims_m.group(0).replace(" ", "").replace("×", "x").replace("X", "x")
        if not dims.endswith("pixels"):
            dims += " pixels"
    if not dims or "x" not in dims.lower():
        dims = "1800x1200 pixels"

    return title, tags, dims, desc


def sanitize_invitation_words(title, tags, desc, subfolder):
    sf = (subfolder or "").lower()
    inv_words = ["invitation", "invitations", "invite", "invites", "nimantrana", "amandhippu"]
    if any(w in sf for w in inv_words):
        return title, tags, desc

    replacements = [
        (r'\bwedding invitation card\b',     'Wedding Banner'),
        (r'\bwedding invitation template\b', 'Wedding Banner Template'),
        (r'\bwedding invitation\b',          'Wedding Banner'),
        (r'\bbirthday invitation card\b',    'Birthday Banner'),
        (r'\bbirthday invitation template\b','Birthday Banner Template'),
        (r'\bbirthday invitation\b',         'Birthday Banner'),
        (r'\bhousewarming invitation\b',     'House Warming Banner'),
        (r'\bhouse warming invitation\b',    'House Warming Banner'),
        (r'\binvitation card psd\b',         'Banner PSD'),
        (r'\binvitation card template\b',    'Banner Template'),
        (r'\binvitation card\b',             'Banner'),
        (r'\binvitation template\b',         'Banner Template'),
        (r'\binvitation psd\b',              'Banner PSD'),
        (r'\binvitation design\b',           'Banner Design'),
        (r'\binvitations\b',                 'Banners'),
        (r'\binvitation\b',                  'Banner'),
        (r'\binvite\b',                      'Banner'),
        (r'\binvites\b',                     'Banners'),
    ]
    for pat, rep in replacements:
        title = re.sub(pat, rep, title, flags=re.IGNORECASE).strip()
        tags  = re.sub(pat, rep, tags,  flags=re.IGNORECASE).strip()
        desc  = re.sub(pat, rep, desc,  flags=re.IGNORECASE).strip()

    # Re-clean tags after replacements
    tags = clean_tags(tags)
    return title, tags, desc


def truncate_description(desc: str, max_words: int = 380, min_words: int = 300) -> str:
    """
    FIX: Previously limited by line count (100 lines), which allowed 2000+ word descriptions.
    Now enforces a strict word count ceiling of max_words (default 380).
    Truncates at sentence boundary when possible.
    """
    if not desc:
        return desc

    words = desc.split()
    if len(words) <= max_words:
        return desc

    # Truncate at word boundary
    truncated = " ".join(words[:max_words])

    # Try to end at a clean sentence boundary (., !, ?)
    last_sentence_end = max(
        truncated.rfind("."),
        truncated.rfind("!"),
        truncated.rfind("?"),
    )
    # Only snap to sentence boundary if it's in the last 20% of text
    if last_sentence_end > int(len(truncated) * 0.80):
        return truncated[:last_sentence_end + 1].strip()

    return truncated.rstrip(".").strip() + "."


# ╔══════════════════════════════════════════════╗
# ║         MODELSCOPE AI CALL                  ║
# ╚══════════════════════════════════════════════╝

def call_modelscope(jpg_b64: str, folder_hint: str = "", retries: int = 5) -> str:
    full_prompt = (folder_hint.strip() + "\n\n" if folder_hint.strip() else "") + SEO_PROMPT
    headers = {
        "Authorization": f"Bearer {MODELSCOPE_TOKEN}",
        "Content-Type":  "application/json",
    }
    payload = {
        "model": VISION_MODEL,
        "messages": [{
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:image/jpeg;base64,{jpg_b64}"}
                },
                {"type": "text", "text": full_prompt}
            ]
        }],
        # FIX: Reduced from 2000 to 900 — enough for title+tags+dims+300-380w description.
        # This prevents the AI from generating 2000+ word descriptions.
        "max_tokens": 900,
        "temperature": 0.75,
    }

    for attempt in range(retries):
        try:
            resp = requests.post(
                MODELSCOPE_API, headers=headers, json=payload, timeout=120
            )

            if resp.status_code == 429:
                wait = 60 * (attempt + 1)
                log(f"    ⚠  ModelScope HTTP 429 rate-limit (attempt {attempt+1}/{retries}) — waiting {wait}s")
                time.sleep(wait)
                continue

            if resp.status_code in (502, 503, 504):
                wait = 30 * (attempt + 1)
                log(f"    ⚠  ModelScope HTTP {resp.status_code} transient error (attempt {attempt+1}/{retries}) — waiting {wait}s")
                time.sleep(wait)
                continue

            data = resp.json()

            if "choices" in data:
                raw_content = data["choices"][0]["message"]["content"].strip()
                log(f"    📝 AI raw (first 200 chars): {raw_content[:200]!r}")
                return raw_content

            err_obj  = data.get("error", {})
            err_code = str(err_obj.get("code", "")).lower() if isinstance(err_obj, dict) else ""
            err_msg  = str(err_obj.get("message", data.get("message", str(data)[:300]))).lower()

            if "quota" in err_msg or "quota" in err_code:
                wait = 90 * (attempt + 1)
                log(f"    ⚠  ModelScope quota error (attempt {attempt+1}/{retries}) — waiting {wait}s: {err_msg[:200]}")
                time.sleep(wait)
                continue

            if "rate" in err_msg or "throttl" in err_msg:
                wait = 60 * (attempt + 1)
                log(f"    ⚠  ModelScope rate-limit error (attempt {attempt+1}/{retries}) — waiting {wait}s")
                time.sleep(wait)
                continue

            log(f"    ⚠  ModelScope no choices (attempt {attempt+1}/{retries}): {str(data)[:300]}")

        except Exception as e:
            log(f"    ⚠  ModelScope request error (attempt {attempt+1}/{retries}): {e}")

        if attempt < retries - 1:
            time.sleep(20 * (attempt + 1))

    return ""


# ╔══════════════════════════════════════════════╗
# ║         TEXT-ONLY FALLBACK                  ║
# ╚══════════════════════════════════════════════╝

def generate_seo_fallback(filename: str, category: str) -> str:
    stem = Path(filename).stem
    prompt = (
        f"You are a Tamil PSD marketplace SEO expert.\n"
        f"Generate SEO content for a Tamil PSD design file.\n"
        f"File name: {stem}\n"
        f"Category folder: {category or 'General'}\n\n"
        + SEO_PROMPT
    )
    headers = {
        "Authorization": f"Bearer {MODELSCOPE_TOKEN}",
        "Content-Type":  "application/json",
    }
    payload = {
        "model": VISION_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 900,
        "temperature": 0.7,
    }
    try:
        resp = requests.post(MODELSCOPE_API, headers=headers, json=payload, timeout=90)
        data = resp.json()
        if "choices" in data:
            return data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        log(f"    ⚠  Text-only fallback also failed: {e}")
    return ""


# ╔══════════════════════════════════════════════╗
# ║         MAIN PIPELINE                       ║
# ╚══════════════════════════════════════════════╝

def main():
    log("\n" + "═" * 70)
    log("🚀  Tamil PSD Marketplace — GitHub Actions Pipeline v2")
    log("═" * 70)

    missing = [k for k in [
        "GOOGLE_CLIENT_ID", "GOOGLE_CLIENT_SECRET", "GOOGLE_REFRESH_TOKEN",
        "GDRIVE_PSD_FOLDER", "GDRIVE_UPLOAD_FOLDER", "SCOPE"
    ] if not os.environ.get(k)]
    if missing:
        log(f"\n❌ Missing required secrets: {', '.join(missing)}")
        log("   Add them to your GitHub Actions repo secrets and retry.")
        sys.exit(1)

    content_token = GH_PAT or GITHUB_TOKEN
    if not content_token:
        log("❌ No GitHub token available. Set GH_PAT secret.")
        sys.exit(1)

    backend_token = GITHUB_TOKEN or GH_PAT

    # Ensure local preview_image/ directory exists
    PREVIEW_LOCAL_DIR.mkdir(parents=True, exist_ok=True)

    # ── STEP 1: Fetch existing designs.xlsx ──────────────────────────────────
    log("\n📥 STEP 1 — Fetching designs.xlsx from Gurumoorthi…")
    df, xlsx_sha = fetch_designs_xlsx(content_token)
    existing_ids = {
        str(r).strip() for r in df["ID"].tolist()
        if str(r).strip() not in ("", "nan", "None")
    }
    log(f"  📊 Existing rows: {len(df)} | Known IDs: {len(existing_ids)}")

    # ── STEP 2: Scan GDrive PSD folder ───────────────────────────────────────
    log(f"\n📂 STEP 2 — Scanning Google Drive folder: {GDRIVE_PSD_FOLDER}")
    all_files = scan_gdrive_structure(GDRIVE_PSD_FOLDER)
    log(f"  📄 Total files in GDrive: {len(all_files)}")

    new_files = [
        (fid, fname, cat, cat_fid) for fid, fname, cat, cat_fid in all_files
        if Path(fname).stem not in existing_ids
    ]
    log(f"  🆕 New files (not yet in xlsx): {len(new_files)}")

    if not new_files:
        log("\n✅ Nothing new to process — designs.xlsx is already up to date!")
        return

    to_process = new_files[:MAX_PER_RUN]
    log(f"  🔢 Processing {len(to_process)} files this run (limit: {MAX_PER_RUN})")

    # ── STEP 3: Process each file ─────────────────────────────────────────────
    log(f"\n🤖 STEP 3 — Download → Convert → Watermark → WebP → Upload → AI → xlsx")
    new_rows       = []
    errors         = 0
    webp_saved_locally = 0

    for idx, (file_id, filename, category, category_folder_id) in enumerate(to_process):
        stem = Path(filename).stem
        ext  = Path(filename).suffix.lower()
        log(f"\n  [{idx+1}/{len(to_process)}]  {filename}  📁[{category or 'root'}]")

        try:
            # ── 3a. Download from GDrive ─────────────────────────────────────
            log(f"    ⬇️  Downloading from GDrive…")
            raw_bytes = gdrive_download(file_id)
            file_size = format_size(len(raw_bytes))
            log(f"    ✅ Downloaded: {file_size}")

            # ── 3b. Convert to WebP preview (with watermark) ─────────────────
            log(f"    🖼️  Converting to WebP (watermarked, <{WEBP_TARGET_KB}KB)…")
            webp_bytes = to_webp_bytes(raw_bytes, filename, max_size=PREVIEW_MAX_SIZE)
            if not webp_bytes:
                log(f"    ❌ WebP conversion failed — skipping {filename}")
                errors += 1
                continue

            # ── 3c. Save WebP: locally first, then GitHub API ─────────────────
            # FIX: Primary save is now LOCAL (committed by git step in workflow).
            # API upload is secondary — if it fails, the local save still works.
            webp_filename = stem + ".webp"
            gh_path       = f"{PREVIEW_DIR}/{webp_filename}"

            # Primary: save to local preview_image/ directory
            saved_locally = save_webp_locally(webp_bytes, webp_filename)
            if saved_locally:
                webp_saved_locally += 1

            # Secondary: also attempt GitHub API upload for immediate availability
            uploaded_ok = github_upload_file(
                BACKEND_REPO, gh_path, webp_bytes, backend_token,
                f"ci: add preview {webp_filename}"
            )

            # Preview URL is valid if saved locally OR API upload succeeded
            preview_url = f"{CDN_BASE}/{webp_filename}" if (saved_locally or uploaded_ok) else ""
            if preview_url:
                log(f"    🔗 jsDelivr: {preview_url}")
            else:
                log(f"    ⚠  WebP not saved — preview URL will be empty")

            # ── 3d. Encode a JPEG version for the AI call ────────────────────
            log(f"    🤖 Preparing JPEG for AI vision call…")
            try:
                if ext in (".psd", ".psb"):
                    from psd_tools import PSDImage
                    psd    = PSDImage.open(io.BytesIO(raw_bytes))
                    ai_img = psd.composite() or PILImage.open(io.BytesIO(raw_bytes))
                else:
                    ai_img = PILImage.open(io.BytesIO(raw_bytes))
                if ai_img.mode not in ("RGB", "RGBA"):
                    ai_img = ai_img.convert("RGB")
                if max(ai_img.size) > 1280:
                    ai_img = ai_img.copy()
                    ai_img.thumbnail((1280, 1280), PILImage.LANCZOS)
                if ai_img.mode == "RGBA":
                    bg = PILImage.new("RGB", ai_img.size, (255, 255, 255))
                    bg.paste(ai_img, mask=ai_img.split()[3])
                    ai_img = bg
                elif ai_img.mode != "RGB":
                    ai_img = ai_img.convert("RGB")
                ai_buf = io.BytesIO()
                ai_img.save(ai_buf, "JPEG", quality=85)
                jpg_b64 = base64.b64encode(ai_buf.getvalue()).decode()
            except Exception as ai_prep_err:
                log(f"    ⚠  AI image prep failed: {ai_prep_err} — using WebP bytes as fallback")
                jpg_b64 = base64.b64encode(webp_bytes).decode()

            # ── 3e. Zip the original file ────────────────────────────────────
            zip_filename = stem + ".zip"
            log(f"    📦 Zipping → {zip_filename}…")
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w",
                                 compression=zipfile.ZIP_DEFLATED,
                                 compresslevel=6,
                                 allowZip64=True) as zf:
                zf.writestr(filename, raw_bytes)
            zip_bytes = zip_buf.getvalue()
            log(f"    ✅ ZIP: {format_size(len(zip_bytes))}")

            # ── 3f. Upload ZIP to GDrive upload folder ───────────────────────
            log(f"    ☁️  Uploading ZIP to GDrive upload folder…")
            dl_url = gdrive_upload(zip_bytes, zip_filename, GDRIVE_UPLOAD_FOLDER)
            if not dl_url:
                log(f"    ⚠  ZIP upload failed — dl_url will be empty")

            # ── 3f2. Move source PSD to 'final' subfolder ────────────────────
            if dl_url and category_folder_id:
                log(f"    📁 Moving source file to final/ subfolder…")
                moved = gdrive_move_to_final(file_id, category_folder_id)
                if moved:
                    log(f"    ✅ Source moved to final/")
                else:
                    log(f"    ⚠  Could not move source to final/ — continuing anyway")
            elif not category_folder_id:
                log(f"    ⚠  No category_folder_id — skipping move to final/")

            # ── 3g. ModelScope vision AI ─────────────────────────────────────
            log(f"    🤖 Running AI vision analysis (ModelScope Qwen2.5-VL)…")
            folder_ctx   = get_folder_context(category)
            folder_hint  = folder_ctx.get("ai_hint", "")
            cat_override = folder_ctx.get("category_hint", "")

            ai_raw = call_modelscope(jpg_b64, folder_hint)

            title = tags = dims = desc = ""
            if ai_raw:
                title, tags, dims, desc = parse_response(ai_raw)

                if not dims or "x" not in dims.lower():
                    dims = "1800x1200 pixels"

                title, tags, desc = sanitize_invitation_words(title, tags, desc, category)
                desc = truncate_description(desc)

                cat_label = cat_override or category or "Others"

                word_count = len(desc.split()) if desc else 0
                log(f"    ✅ Title    : {title[:70]}")
                log(f"    ✅ Category : {cat_label}")
                log(f"    ✅ Tags     : {tags[:70]}…")
                log(f"    ✅ Dims     : {dims}  |  Desc: {word_count}w")
                log(f"    🔗 DL: {'✅' if dl_url else '❌'}  Preview: {'✅' if preview_url else '❌'}")
            else:
                log(f"    ⚠  AI vision returned empty — trying text-only fallback...")
                ai_raw = generate_seo_fallback(filename, category)
                if ai_raw:
                    title, tags, dims, desc = parse_response(ai_raw)
                    if not dims or "x" not in dims.lower():
                        dims = "1800x1200 pixels"
                    title, tags, desc = sanitize_invitation_words(title, tags, desc, category)
                    desc = truncate_description(desc)
                    cat_label = cat_override or category or "Others"
                    log(f"    ✅ Fallback Title: {title[:70]}")
                else:
                    log(f"    ❌ Both vision and fallback failed — SEO fields will be empty")
                    cat_label = category or "Others"

            color_mode = "CMYK" if ext in (".psd", ".psb") else "RGB"

            new_rows.append([
                stem,                       # ID
                dl_url,                     # Download URL
                title,                      # Title
                cat_label,                  # Category
                tags,                       # Tags
                desc,                       # Description
                dims,                       # Dimensions
                "300 DPI",                  # DPI
                file_size,                  # File Size
                color_mode,                 # Color Mode
                "Adobe Photoshop CC",       # Software
                "",                         # Fonts Used
                preview_url,                # Preview URL
            ])

            time.sleep(AI_DELAY)

        except Exception as e:
            log(f"    ❌ Unhandled error for {filename}: {e}")
            traceback.print_exc()
            errors += 1
            continue

    # ── STEP 4: Merge new rows ────────────────────────────────────────────────
    log(f"\n🔀 STEP 4 — Merging {len(new_rows)} new rows into designs.xlsx…")
    if new_rows:
        new_df   = pd.DataFrame(new_rows, columns=XLSX_HEADERS)
        final_df = pd.concat([df, new_df], ignore_index=True)
    else:
        final_df = df.copy()

    # ── STEP 5: Push designs.xlsx to Gurumoorthi ─────────────────────────────
    log(f"\n🚀 STEP 5 — Pushing designs.xlsx to {CONTENT_REPO}…")
    commit_msg = (
        f"ci: add {len(new_rows)} designs [{time.strftime('%Y-%m-%d %H:%M UTC', time.gmtime())}]"
    )
    push_designs_xlsx(final_df, xlsx_sha, content_token, commit_msg)

    log("\n" + "═" * 70)
    log("✅  PIPELINE COMPLETE")
    log(f"    New files processed    : {len(new_rows)}")
    log(f"    WebP saved locally     : {webp_saved_locally}  ← committed by git step")
    log(f"    Errors                 : {errors}")
    log(f"    Total rows in xlsx     : {len(final_df)}")
    log("═" * 70)

    if errors > 0 and len(new_rows) == 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
