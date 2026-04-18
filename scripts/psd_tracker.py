"""
psd_tracker.py – Single Excel file as the COMPLETE source of truth.

This replaces BOTH state_manager.py AND excel_tracker.py with one clean system.

Excel file: psd_tracker.xlsx  (saved locally + pushed to GitHub)

Sheet "Pipeline" columns:
  A  url              – forpsd.com download URL (unique key)
  B  detail_url       – product detail page URL
  C  card_title       – title text from listing card (for category hint)
  D  category         – resolved GDrive subfolder name
  E  status           – pending | completed | error | skipped
  F  original_files   – comma-joined list of original filenames inside archive
  G  renamed_files    – comma-joined list of tamilpsd-XXXX names assigned
  H  drive_links      – comma-joined list of Google Drive links
  I  scraped_at       – ISO datetime when URL was discovered
  J  processed_at     – ISO datetime when upload completed
  K  error_msg        – error details if status=error

Usage:
    tracker = PSDTracker("psd_tracker.xlsx", github_token, "owner/repo")
    tracker.load_from_github()           # fetch latest from GitHub on startup
    tracker.add_items(scraped_items)     # add new URLs (duplicates skipped)
    pending = tracker.get_pending()      # list of dicts
    tracker.mark_completed(url, ...)     # update row + save + push
    tracker.mark_error(url, msg)         # update row + save + push
"""

import io
import json
import logging
import re
import time
from datetime import datetime, timezone
from pathlib import Path

import base64
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

log = logging.getLogger(__name__)

# ── Column indices (1-based for openpyxl) ─────────────────────────────────────
_C = {
    "url":            1,
    "detail_url":     2,
    "card_title":     3,
    "category":       4,
    "status":         5,
    "original_files": 6,
    "renamed_files":  7,
    "drive_links":    8,
    "scraped_at":     9,
    "processed_at":   10,
    "error_msg":      11,
}
_HEADERS = list(_C.keys())
_NCOLS   = len(_HEADERS)

# ── Styling ───────────────────────────────────────────────────────────────────
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_STATUS_FILLS = {
    "pending":   PatternFill("solid", fgColor="FFF2CC"),  # yellow
    "completed": PatternFill("solid", fgColor="E2EFDA"),  # green
    "error":     PatternFill("solid", fgColor="FCE4D6"),  # red/orange
    "skipped":   PatternFill("solid", fgColor="EDEDED"),  # grey
}
_DEFAULT_FILL_A = PatternFill("solid", fgColor="FFFFFF")
_DEFAULT_FILL_B = PatternFill("solid", fgColor="F5F5F5")

_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right= Side(style="thin", color="CCCCCC"),
)

_COL_WIDTHS = {
    "url": 60, "detail_url": 50, "card_title": 40, "category": 22,
    "status": 12, "original_files": 45, "renamed_files": 35,
    "drive_links": 70, "scraped_at": 20, "processed_at": 20, "error_msg": 50,
}

_TAMILPSD_RE = re.compile(r"tamilpsd-(\d+)\.", re.IGNORECASE)


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


class PSDTracker:
    """
    Single-file Excel tracker for the entire PSD pipeline.

    Internal state:
      _rows        : list of dicts (one per URL), in-order
      _url_index   : dict url → row_index (for O(1) lookup)
      _originals   : set of lowercase original filenames (skip duplicates)
      _renamed     : set of lowercase renamed filenames (avoid counter collisions)
      max_counter  : highest tamilpsd-N number seen (for file counter sync)
    """

    SHEET_NAME   = "Pipeline"
    FILE_NAME    = "psd_tracker.xlsx"

    def __init__(
        self,
        local_path: str,
        github_token: str,
        github_repo: str,
        github_branch: str = "main",
    ):
        self._path         = Path(local_path)
        self._gh_token     = github_token
        self._gh_repo      = github_repo
        self._gh_branch    = github_branch
        self._gh_file_path = self.FILE_NAME   # path inside repo

        # In-memory state
        self._rows:      list[dict] = []
        self._url_index: dict       = {}   # url → list index
        self._originals: set        = set()
        self._renamed:   set        = set()
        self.max_counter: int       = 0

        # SHA needed for GitHub push
        self._gh_sha = None

    # ══════════════════════════════════════════════════════════════════════
    #  GITHUB I/O
    # ══════════════════════════════════════════════════════════════════════

    def _gh_headers(self) -> dict:
        return {
            "Authorization": f"token {self._gh_token}",
            "Accept":        "application/vnd.github.v3+json",
        }

    def load_from_github(self) -> bool:
        """
        Fetch psd_tracker.xlsx from GitHub → parse → populate in-memory state.
        Falls back to empty state if file doesn't exist yet (first run).
        Returns True if loaded, False if starting fresh.
        """
        url = f"https://api.github.com/repos/{self._gh_repo}/contents/{self._gh_file_path}"
        try:
            resp = requests.get(url, headers=self._gh_headers(), timeout=30)
        except Exception as e:
            log.warning(f"GitHub fetch error: {e} — starting fresh")
            return False

        if resp.status_code == 404:
            log.info("📊 psd_tracker.xlsx not found on GitHub — starting fresh (first run)")
            return False

        if resp.status_code != 200:
            log.warning(f"GitHub fetch HTTP {resp.status_code} — starting fresh")
            return False

        data = resp.json()
        self._gh_sha = data.get("sha")

        raw_b64 = data.get("content", "")
        if not raw_b64:
            # Large file — use download_url
            dl_url = data.get("download_url", "")
            if dl_url:
                try:
                    r2 = requests.get(dl_url, headers=self._gh_headers(), timeout=120)
                    raw = r2.content
                except Exception as e:
                    log.warning(f"Download URL fetch error: {e} — starting fresh")
                    return False
            else:
                log.warning("No content or download_url — starting fresh")
                return False
        else:
            try:
                raw = base64.b64decode(raw_b64.replace("\n", ""))
            except Exception as e:
                log.warning(f"base64 decode error: {e} — starting fresh")
                return False

        # Validate ZIP magic bytes
        if raw[:4] != b"PK\x03\x04":
            log.warning("psd_tracker.xlsx is not a valid xlsx — starting fresh")
            return False

        try:
            self._parse_xlsx(raw)
            log.info(
                f"✅ psd_tracker.xlsx loaded from GitHub: "
                f"{len(self._rows)} rows | "
                f"{self._count_by_status('pending')} pending | "
                f"{self._count_by_status('completed')} completed | "
                f"max counter: tamilpsd-{self.max_counter:04d}"
            )
            return True
        except Exception as e:
            log.error(f"xlsx parse error: {e} — starting fresh")
            return False

    def push_to_github(self, commit_msg: str = "") -> bool:
        """
        Save current state to disk AND push to GitHub.
        Uses SHA for update; no SHA = create new file.
        Returns True on success.
        """
        # 1. Save to local disk first
        xlsx_bytes = self._to_xlsx_bytes()
        self._path.parent.mkdir(parents=True, exist_ok=True)
        with open(self._path, "wb") as f:
            f.write(xlsx_bytes)

        # 2. Push to GitHub
        if not commit_msg:
            pending   = self._count_by_status("pending")
            completed = self._count_by_status("completed")
            commit_msg = (
                f"tracker: {completed} completed | {pending} pending "
                f"[{_now_iso()}]"
            )

        url = (
            f"https://api.github.com/repos/{self._gh_repo}"
            f"/contents/{self._gh_file_path}"
        )
        payload: dict = {
            "message": commit_msg,
            "content": base64.b64encode(xlsx_bytes).decode(),
            "branch":  self._gh_branch,
        }
        if self._gh_sha:
            payload["sha"] = self._gh_sha

        try:
            resp = requests.put(
                url, headers=self._gh_headers(), json=payload, timeout=60
            )
        except Exception as e:
            log.error(f"GitHub push error: {e}")
            return False

        if resp.status_code in (200, 201):
            self._gh_sha = resp.json().get("content", {}).get("sha", self._gh_sha)
            log.info(f"✅ psd_tracker.xlsx pushed to GitHub  SHA={str(self._gh_sha)[:10]}…")
            return True

        log.error(f"GitHub push failed HTTP {resp.status_code}: {resp.text[:300]}")
        return False

    # ══════════════════════════════════════════════════════════════════════
    #  XLSX PARSE / BUILD
    # ══════════════════════════════════════════════════════════════════════

    def _parse_xlsx(self, raw_bytes: bytes) -> None:
        """Parse xlsx bytes → populate _rows, _url_index, _originals, _renamed."""
        self._rows      = []
        self._url_index = {}
        self._originals = set()
        self._renamed   = set()
        self.max_counter = 0

        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb[self.SHEET_NAME] if self.SHEET_NAME in wb.sheetnames else wb.active

        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row or not row[0]:
                continue
            record = {
                col: (str(row[i]).strip() if row[i] not in (None, "") else "")
                for i, col in enumerate(_HEADERS)
                if i < len(row)
            }
            # Fill any missing columns
            for col in _HEADERS:
                record.setdefault(col, "")

            url = record.get("url", "")
            if not url:
                continue

            idx = len(self._rows)
            self._rows.append(record)
            self._url_index[url] = idx

            # Build original/renamed sets from completed rows
            if record.get("status") == "completed":
                for fn in _split_csv(record.get("original_files", "")):
                    self._originals.add(fn.lower())
                for rn in _split_csv(record.get("renamed_files", "")):
                    self._renamed.add(rn.lower())
                    m = _TAMILPSD_RE.search(rn)
                    if m:
                        self.max_counter = max(self.max_counter, int(m.group(1)))

        wb.close()

    def _to_xlsx_bytes(self) -> bytes:
        """Render current _rows to styled xlsx bytes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.SHEET_NAME

        # ── Header row ──────────────────────────────────────────────────────
        for ci, col in enumerate(_HEADERS, 1):
            cell = ws.cell(1, ci, col.replace("_", " ").title())
            cell.font      = _HDR_FONT
            cell.fill      = _HDR_FILL
            cell.alignment = _HDR_ALIGN
        ws.row_dimensions[1].height = 28

        # ── Column widths + freeze ───────────────────────────────────────────
        for ci, col in enumerate(_HEADERS, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(ci)
            ].width = _COL_WIDTHS.get(col, 20)
        ws.freeze_panes = "A2"

        # ── Data rows ────────────────────────────────────────────────────────
        for ri, record in enumerate(self._rows, 2):
            status = record.get("status", "pending")
            status_fill = _STATUS_FILLS.get(status)
            alt_fill    = _DEFAULT_FILL_B if ri % 2 == 0 else _DEFAULT_FILL_A

            for ci, col in enumerate(_HEADERS, 1):
                cell = ws.cell(ri, ci, record.get(col, ""))
                cell.border    = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=(col in ("error_msg", "drive_links")))
                # Status cell gets colour; other cells use alt-row shading
                if col == "status" and status_fill:
                    cell.fill = status_fill
                elif col == "status":
                    cell.fill = alt_fill
                else:
                    cell.fill = alt_fill

        # Auto-filter on header
        ws.auto_filter.ref = (
            f"A1:{openpyxl.utils.get_column_letter(_NCOLS)}1"
        )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ══════════════════════════════════════════════════════════════════════
    #  PUBLIC API — QUERY
    # ══════════════════════════════════════════════════════════════════════

    def _count_by_status(self, status: str) -> int:
        return sum(1 for r in self._rows if r.get("status") == status)

    def stats(self) -> str:
        total     = len(self._rows)
        pending   = self._count_by_status("pending")
        completed = self._count_by_status("completed")
        error     = self._count_by_status("error")
        skipped   = self._count_by_status("skipped")
        return (
            f"Total={total} | Pending={pending} | Completed={completed} | "
            f"Error={error} | Skipped={skipped} | "
            f"Next file=tamilpsd-{self.max_counter + 1:04d}"
        )

    def get_pending(self) -> list[dict]:
        """Return all rows with status='pending', in order scraped."""
        return [r for r in self._rows if r.get("status") == "pending"]

    def total_rows(self) -> int:
        return len(self._rows)

    def all_urls(self) -> set:
        return set(self._url_index.keys())

    def is_original_done(self, original_filename: str) -> bool:
        """True if this original file was already uploaded (skip check)."""
        return original_filename.strip().lower() in self._originals

    def is_renamed_used(self, renamed_filename: str) -> bool:
        """True if this tamilpsd-XXXX name is already taken."""
        return renamed_filename.strip().lower() in self._renamed

    # ══════════════════════════════════════════════════════════════════════
    #  PUBLIC API — WRITE
    # ══════════════════════════════════════════════════════════════════════

    def add_items(self, items: list[dict], push: bool = False) -> int:
        """
        Add new items (from scraper) to the tracker.
        Each item must have: download_url, detail_url, card_title.
        Items whose URL already exists in the tracker are silently skipped.

        Returns count of newly added rows.
        """
        added = 0
        now   = _now_iso()
        for item in items:
            url = item.get("download_url", "").strip()
            if not url or url in self._url_index:
                continue
            record = {
                "url":            url,
                "detail_url":     item.get("detail_url", "").strip(),
                "card_title":     item.get("card_title", "").strip(),
                "category":       "",          # resolved later, during processing
                "status":         "pending",
                "original_files": "",
                "renamed_files":  "",
                "drive_links":    "",
                "scraped_at":     now,
                "processed_at":   "",
                "error_msg":      "",
            }
            self._url_index[url] = len(self._rows)
            self._rows.append(record)
            added += 1

        if added > 0:
            log.info(f"📋 Tracker: added {added} new pending items (total: {len(self._rows)})")
            if push:
                self.push_to_github(
                    f"tracker: +{added} new links scraped [{_now_iso()}]"
                )
        return added

    def update_category(self, url: str, category: str) -> None:
        """Set the resolved category for a pending row (no push — done inline)."""
        if url in self._url_index:
            self._rows[self._url_index[url]]["category"] = category

    def mark_completed(
        self,
        url: str,
        original_files: list[str],
        renamed_files:  list[str],
        drive_links:    list[str],
        push: bool = True,
    ) -> bool:
        """
        Mark a URL as completed, record files, and push to GitHub.
        Returns True on successful push.
        """
        if url not in self._url_index:
            log.warning(f"mark_completed: URL not found in tracker: {url[:60]}")
            return False

        record = self._rows[self._url_index[url]]
        record["status"]         = "completed"
        record["original_files"] = ", ".join(original_files)
        record["renamed_files"]  = ", ".join(renamed_files)
        record["drive_links"]    = ", ".join(drive_links)
        record["processed_at"]   = _now_iso()
        record["error_msg"]      = ""

        # Update in-memory skip sets
        for fn in original_files:
            self._originals.add(fn.lower())
        for rn in renamed_files:
            self._renamed.add(rn.lower())
            m = _TAMILPSD_RE.search(rn)
            if m:
                self.max_counter = max(self.max_counter, int(m.group(1)))

        log.info(
            f"✅ Completed: {url[:60]}  "
            f"→ {', '.join(renamed_files)}  "
            f"({len(renamed_files)} file(s))"
        )

        if push:
            return self.push_to_github()
        return True

    def mark_error(self, url: str, error_msg: str, push: bool = True) -> bool:
        """Mark a URL as error. Error rows will be retried on next run."""
        if url not in self._url_index:
            log.warning(f"mark_error: URL not found in tracker: {url[:60]}")
            return False

        record = self._rows[self._url_index[url]]
        record["status"]       = "error"
        record["processed_at"] = _now_iso()
        record["error_msg"]    = error_msg[:500]

        log.warning(f"❌ Error marked: {url[:60]}  → {error_msg[:100]}")

        if push:
            return self.push_to_github()
        return True

    def reset_errors_to_pending(self) -> int:
        """
        Reset all error rows back to 'pending' so they are retried next run.
        Call this at the start of a run if desired.
        Returns count reset.
        """
        reset = 0
        for record in self._rows:
            if record.get("status") == "error":
                record["status"]    = "pending"
                record["error_msg"] = ""
                reset += 1
        if reset:
            log.info(f"🔄 Reset {reset} error rows → pending (will retry)")
        return reset


# ── Helpers ────────────────────────────────────────────────────────────────────

def _split_csv(value: str) -> list[str]:
    """Split comma-separated string, stripping whitespace and empty strings."""
    if not value:
        return []
    return [v.strip() for v in value.split(",") if v.strip()]
