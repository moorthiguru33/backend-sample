"""
job_tracker.py – Excel-based primary job tracker for PSD Automation Pipeline.

┌─────────────────────────────────────────────────────────────────────┐
│  SOURCE OF TRUTH: jobs.xlsx                                         │
│                                                                     │
│  First run  → Full scrape → add ALL URLs as Pending (batch write)   │
│  Every run  → Read Pending rows → Process → mark_completed          │
│  All done   → Re-scrape from page 1 → add new URLs → continue      │
└─────────────────────────────────────────────────────────────────────┘

PERFORMANCE CRITICAL:
  add_pending_items(6918 items) → ONE single workbook save (not 6918!)
  mark_completed()              → targeted cell update (not full rewrite!)

Excel columns (jobs.xlsx):
  A  Download URL  – the /download/eyJ… URL from forpsd.com
  B  Detail URL    – product page URL (for category fallback)
  C  Card Title    – title from listing card (fast category detection)
  D  Category      – folder category (resolved; empty until first process)
  E  Status        – Pending / Completed / Error
  F  Date Added    – UTC timestamp when first discovered
  G  Date Completed – UTC timestamp when upload succeeded
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

log = logging.getLogger(__name__)

# ── Column indices (1-based) ───────────────────────────────────────────────
_C_URL       = 1   # A: Download URL
_C_DETAIL    = 2   # B: Detail URL
_C_TITLE     = 3   # C: Card Title
_C_CATEGORY  = 4   # D: Category
_C_STATUS    = 5   # E: Status
_C_ADDED     = 6   # F: Date Added
_C_COMPLETED = 7   # G: Date Completed

STATUS_PENDING   = "Pending"
STATUS_COMPLETED = "Completed"
STATUS_ERROR     = "Error"

_HEADERS    = [
    "Download URL", "Detail URL", "Card Title",
    "Category", "Status", "Date Added", "Date Completed"
]
_COL_WIDTHS = [80, 55, 40, 22, 14, 22, 22]

# ── Styles ─────────────────────────────────────────────────────────────────
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
_PENDING_ODD  = PatternFill("solid", fgColor="FFFFF0")   # ivory
_PENDING_EVEN = PatternFill("solid", fgColor="FFF9E6")   # light yellow
_DONE_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green
_ERROR_FILL   = PatternFill("solid", fgColor="FCE4D6")   # light red
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP         = Alignment(wrap_text=True, vertical="top")
_THIN_BORDER  = Border(
    bottom=Side(style="thin", color="D0D0D0"),
    right =Side(style="thin", color="D0D0D0"),
)


def _col_letter(col: int) -> str:
    result = ""
    while col:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _row_fill(status: str, row_idx: int) -> PatternFill:
    if status == STATUS_COMPLETED:
        return _DONE_FILL
    if status == STATUS_ERROR:
        return _ERROR_FILL
    return _PENDING_EVEN if row_idx % 2 == 0 else _PENDING_ODD


class JobTracker:
    """
    Excel-based URL job tracker. O(1) lookups via in-memory dict.

    ⚡ Performance notes:
      • add_pending_items(N)  → ONE xlsx open+save regardless of N
      • mark_completed(url)   → ONE xlsx open + targeted cell update + save
      • Never does O(N) saves for O(N) items
    """

    def __init__(self, xlsx_path: str | Path):
        self._path      = Path(xlsx_path)
        # url → {detail, card_title, category, status, date_added, date_completed, row}
        self._jobs: dict[str, dict] = {}
        self._next_row: int = 2          # row 1 = header
        self._load()

    # ── Private: Load ──────────────────────────────────────────────────────

    def _load(self) -> None:
        if not self._path.exists():
            log.info("📋 Job tracker: no jobs.xlsx yet — fresh start.")
            return
        try:
            wb = openpyxl.load_workbook(self._path)
            ws = wb.active
            counts = {STATUS_PENDING: 0, STATUS_COMPLETED: 0, STATUS_ERROR: 0}
            row_num = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                url = str(row[0]).strip() if row[0] else ""
                if not url or url in ("None", "nan", "Download URL"):
                    continue

                # Read all columns (handle old format without card_title)
                if len(row) >= 7:
                    detail    = str(row[1]).strip() if row[1] else ""
                    card_title= str(row[2]).strip() if row[2] else ""
                    category  = str(row[3]).strip() if row[3] else ""
                    status    = str(row[4]).strip() if row[4] else STATUS_PENDING
                    added     = str(row[5]).strip() if row[5] else ""
                    completed = str(row[6]).strip() if row[6] else ""
                elif len(row) >= 6:
                    # Old format (no card_title): URL, Detail, Category, Status, Added, Done
                    detail    = str(row[1]).strip() if row[1] else ""
                    card_title= ""
                    category  = str(row[2]).strip() if row[2] else ""
                    status    = str(row[3]).strip() if row[3] else STATUS_PENDING
                    added     = str(row[4]).strip() if row[4] else ""
                    completed = str(row[5]).strip() if row[5] else ""
                else:
                    detail = card_title = category = added = completed = ""
                    status = STATUS_PENDING

                if status not in (STATUS_PENDING, STATUS_COMPLETED, STATUS_ERROR):
                    status = STATUS_PENDING

                self._jobs[url] = {
                    "detail": detail, "card_title": card_title,
                    "category": category, "status": status,
                    "date_added": added, "date_completed": completed,
                    "row": row_num,
                }
                counts[status] = counts.get(status, 0) + 1

            self._next_row = row_num + 1
            log.info(
                f"📋 Job tracker loaded: {len(self._jobs)} total | "
                f"Pending={counts[STATUS_PENDING]} "
                f"Completed={counts[STATUS_COMPLETED]} "
                f"Error={counts[STATUS_ERROR]}"
            )
        except Exception as exc:
            log.error(f"Job tracker load error: {exc} — starting fresh.")
            self._jobs     = {}
            self._next_row = 2

    # ── Private: Workbook helpers ──────────────────────────────────────────

    def _make_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"
        for ci, (hdr, w) in enumerate(zip(_HEADERS, _COL_WIDTHS), 1):
            cell           = ws.cell(1, ci, hdr)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _CENTER
            ws.column_dimensions[_col_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{_col_letter(len(_HEADERS))}1"
        return wb

    def _load_or_create_wb(self) -> openpyxl.Workbook:
        self._path.parent.mkdir(parents=True, exist_ok=True)
        if self._path.exists():
            try:
                return openpyxl.load_workbook(self._path)
            except Exception:
                log.warning("jobs.xlsx corrupted — rebuilding.")
        return self._make_workbook()

    def _write_row(self, ws, ri: int, url: str, info: dict) -> None:
        """Write one data row into the worksheet."""
        fill   = _row_fill(info.get("status", STATUS_PENDING), ri)
        values = [
            url,
            info.get("detail", ""),
            info.get("card_title", ""),
            info.get("category", ""),
            info.get("status", STATUS_PENDING),
            info.get("date_added", ""),
            info.get("date_completed", ""),
        ]
        for ci, val in enumerate(values, 1):
            cell           = ws.cell(ri, ci, val or "")
            cell.fill      = fill
            cell.border    = _THIN_BORDER
            cell.alignment = _WRAP

    # ── Public API ─────────────────────────────────────────────────────────

    def has_url(self, url: str) -> bool:
        return url.strip() in self._jobs

    def is_completed(self, url: str) -> bool:
        return self._jobs.get(url.strip(), {}).get("status") == STATUS_COMPLETED

    def total(self) -> int:
        return len(self._jobs)

    def completed_count(self) -> int:
        return sum(1 for v in self._jobs.values() if v.get("status") == STATUS_COMPLETED)

    def get_pending(self) -> list[dict]:
        """
        Return list of pending/error items in insertion order.
        Each dict: {download_url, detail_url, card_title, category}
        """
        result = []
        for url, info in sorted(self._jobs.items(), key=lambda x: x[1]["row"]):
            if info.get("status") in (STATUS_PENDING, STATUS_ERROR):
                result.append({
                    "download_url": url,
                    "detail_url":   info.get("detail", ""),
                    "card_title":   info.get("card_title", ""),
                    "category":     info.get("category", ""),
                })
        return result

    def all_known_urls(self) -> set[str]:
        return set(self._jobs.keys())

    def add_pending_items(self, items: list[dict]) -> int:
        """
        ⚡ Add new items with Status=Pending using a SINGLE xlsx save.
        Items already tracked (any status) are silently skipped.
        Returns count of NEW items added.
        """
        now      = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        new_urls = []

        for item in items:
            url = item.get("download_url", "").strip()
            if not url or url in self._jobs:
                continue
            info = {
                "detail":         item.get("detail_url", ""),
                "card_title":     item.get("card_title", ""),
                "category":       item.get("category", ""),
                "status":         STATUS_PENDING,
                "date_added":     now,
                "date_completed": "",
                "row":            self._next_row,
            }
            self._jobs[url] = info
            new_urls.append(url)
            self._next_row += 1

        if not new_urls:
            return 0

        # ── Single workbook open + write all rows + save ──────────────────
        try:
            wb = self._load_or_create_wb()
            ws = wb.active
            for url in new_urls:
                info = self._jobs[url]
                self._write_row(ws, info["row"], url, info)
            wb.save(self._path)
            log.info(f"📋 Batch saved {len(new_urls)} new Pending rows → {self._path.name}")
        except Exception as exc:
            log.error(f"add_pending_items batch save error: {exc}")

        return len(new_urls)

    def update_category(self, url: str, category: str) -> None:
        """Cache resolved category in memory (persisted on next mark_completed)."""
        url = url.strip()
        if url in self._jobs and category:
            self._jobs[url]["category"] = category

    def mark_completed(self, url: str) -> None:
        """
        ⚡ Update Status=Completed using targeted cell writes (no full rewrite).
        """
        url = url.strip()
        if url not in self._jobs:
            log.warning(f"mark_completed: unknown URL {url[:60]}")
            return
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        self._jobs[url]["status"]         = STATUS_COMPLETED
        self._jobs[url]["date_completed"] = now

        try:
            wb  = self._load_or_create_wb()
            ws  = wb.active
            ri  = self._jobs[url]["row"]
            cat = self._jobs[url].get("category", "")
            # Update all 7 cells in this row (ensures color + category are saved)
            values = [
                url,
                self._jobs[url].get("detail", ""),
                self._jobs[url].get("card_title", ""),
                cat,
                STATUS_COMPLETED,
                self._jobs[url].get("date_added", ""),
                now,
            ]
            for ci, val in enumerate(values, 1):
                cell           = ws.cell(ri, ci, val or "")
                cell.fill      = _DONE_FILL
                cell.border    = _THIN_BORDER
                cell.alignment = _WRAP
            wb.save(self._path)
            log.info(f"📋 ✅ Completed [{ri}]: {url[:65]}")
        except Exception as exc:
            log.error(f"mark_completed save error: {exc}")

    def mark_error(self, url: str) -> None:
        """Mark URL as Error (will be retried next run) with targeted cell update."""
        url = url.strip()
        if url not in self._jobs:
            return
        self._jobs[url]["status"] = STATUS_ERROR
        try:
            wb = self._load_or_create_wb()
            ws = wb.active
            ri = self._jobs[url]["row"]
            ws.cell(ri, _C_STATUS, STATUS_ERROR).fill = _ERROR_FILL
            for ci in range(1, len(_HEADERS) + 1):
                ws.cell(ri, ci).fill = _ERROR_FILL
            wb.save(self._path)
            log.warning(f"📋 ❌ Error [{ri}]: {url[:65]}")
        except Exception as exc:
            log.error(f"mark_error save error: {exc}")

    def stats(self) -> str:
        total = len(self._jobs)
        done  = sum(1 for v in self._jobs.values() if v.get("status") == STATUS_COMPLETED)
        err   = sum(1 for v in self._jobs.values() if v.get("status") == STATUS_ERROR)
        pend  = total - done - err
        return f"Total={total} | Pending={pend} | Completed={done} | Error={err}"
