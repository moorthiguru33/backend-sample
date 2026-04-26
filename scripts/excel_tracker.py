"""
excel_tracker.py – Persistent Excel log of every file rename + upload.

Columns
-------
  A  Original File Name  – filename as it was inside the source archive
  B  Renamed File Name   – the tamilpsd-XXXX.ext name assigned on upload

Purpose
-------
  • Skip already-done files across runs without touching Google Drive.
  • Guarantee no duplicate renamed names (tamilpsd-XXXX) are ever reused.
  • Guarantee no original file is processed twice (even if it appears in a
    re-downloaded archive or a different archive on a later run).
  • Single source of truth: on every run we load this file first and derive
    the correct starting file_counter from it, so state.json is secondary.

Usage
-----
    tracker = ExcelTracker("data/rename_log.xlsx")

    # Before processing a file from an archive:
    if tracker.is_original_done("my_design.psd"):
        skip ...

    # Before uploading (double-check renamed name is free):
    if tracker.is_renamed_used("tamilpsd-0042.psd"):
        skip ...

    # After a successful upload:
    tracker.add_entry("my_design.psd", "tamilpsd-0042.psd")
    # ↑ saves to disk immediately; no data loss if the run crashes later.

    # At startup — sync file counter:
    start_counter = max(state_counter, drive_counter, tracker.max_counter + 1)
"""

import logging
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

_COL_ORIGINAL = 1   # Column A
_COL_RENAMED  = 2   # Column B
_COL_URL      = 3   # Column C
_HEADER_ROW   = 1

_TAMILPSD_RE = re.compile(r"tamilpsd-(\d+)\.", re.IGNORECASE)

# ── Styling constants ──────────────────────────────────────────────────────
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="2F5496")   # dark blue
_ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")   # light blue
_CENTER       = Alignment(horizontal="center", vertical="center")


class ExcelTracker:
    """
    Persistent three-column Excel log.

    All lookups are O(1) via in-memory sets; the file is only read once
    (at __init__) and written once per successful upload (add_entry).
    """

    def __init__(self, xlsx_path: str | Path):
        self._path: Path = Path(xlsx_path)

        # In-memory sets for instant duplicate checks
        self._originals: set[str] = set()   # lowercase original filenames
        self._renamed:   set[str] = set()   # lowercase renamed filenames
        self._urls:      set[str] = set()   # lowercase source URLs

        # Highest tamilpsd-N number seen in the sheet
        self.max_counter: int = 0

        self._load()

    # ── Private ────────────────────────────────────────────────────────────

    def _load(self) -> None:
        """Read existing file and populate in-memory sets."""
        if not self._path.exists():
            log.info(f"📊 Excel tracker: no file at {self._path} — will create on first entry.")
            return

        try:
            wb = openpyxl.load_workbook(self._path)
            ws = wb.active

            loaded = 0
            for row in ws.iter_rows(min_row=_HEADER_ROW + 1, values_only=True):
                orig    = str(row[0]).strip() if row[0] else ""
                renamed = str(row[1]).strip() if row[1] else ""
                url     = str(row[2]).strip() if (len(row) >= 3 and row[2]) else ""

                if not orig and not renamed:
                    continue

                if orig:
                    self._originals.add(orig.lower())
                if renamed:
                    self._renamed.add(renamed.lower())
                    m = _TAMILPSD_RE.search(renamed)
                    if m:
                        self.max_counter = max(self.max_counter, int(m.group(1)))
                if url:
                    self._urls.add(url.lower())

                loaded += 1

            log.info(
                f"📊 Excel tracker loaded: {loaded} entries from {self._path} | "
                f"Highest renamed counter: tamilpsd-{self.max_counter:04d}"
            )
        except Exception as exc:
            log.error(f"Excel tracker load error ({self._path}): {exc} — starting fresh.")

    def _create_workbook(self) -> openpyxl.Workbook:
        """Create a brand-new workbook with styled headers."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rename Log"

        # Headers
        ws.cell(_HEADER_ROW, _COL_ORIGINAL, "Original File Name").font  = _HEADER_FONT
        ws.cell(_HEADER_ROW, _COL_ORIGINAL).fill                         = _HEADER_FILL
        ws.cell(_HEADER_ROW, _COL_ORIGINAL).alignment                    = _CENTER
        
        ws.cell(_HEADER_ROW, _COL_RENAMED,  "Renamed File Name").font   = _HEADER_FONT
        ws.cell(_HEADER_ROW, _COL_RENAMED).fill                          = _HEADER_FILL
        ws.cell(_HEADER_ROW, _COL_RENAMED).alignment                     = _CENTER

        ws.cell(_HEADER_ROW, _COL_URL,      "Source URL").font         = _HEADER_FONT
        ws.cell(_HEADER_ROW, _COL_URL).fill                             = _HEADER_FILL
        ws.cell(_HEADER_ROW, _COL_URL).alignment                        = _CENTER

        # Column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 70

        # Freeze header row
        ws.freeze_panes = "A2"

        return wb

    def _save(self, original: str, renamed: str, url: str) -> None:
        """Append one row to the Excel file and save."""
        self._path.parent.mkdir(parents=True, exist_ok=True)

        if self._path.exists():
            try:
                wb = openpyxl.load_workbook(self._path)
            except Exception:
                log.warning("Excel tracker: corrupted file — rebuilding.")
                wb = self._create_workbook()
        else:
            wb = self._create_workbook()

        ws = wb.active

        # Next empty row
        next_row = ws.max_row + 1

        # Alternate row colour for readability
        fill = _ALT_FILL if next_row % 2 == 0 else None

        orig_cell    = ws.cell(next_row, _COL_ORIGINAL, original)
        renamed_cell = ws.cell(next_row, _COL_RENAMED,  renamed)
        url_cell     = ws.cell(next_row, _COL_URL,      url)

        if fill:
            orig_cell.fill    = fill
            renamed_cell.fill = fill
            url_cell.fill     = fill

        try:
            wb.save(self._path)
        except Exception as exc:
            log.error(f"Excel tracker save error: {exc}")

    # ── Public API ─────────────────────────────────────────────────────────

    def is_original_done(self, original_filename: str) -> bool:
        """
        Return True if this original filename has already been processed
        and uploaded in a previous run.
        """
        return original_filename.strip().lower() in self._originals

    def is_renamed_used(self, renamed_filename: str) -> bool:
        """
        Return True if this renamed filename (tamilpsd-XXXX.ext) already
        exists in the log — prevents counter collisions and duplicates.
        """
        return renamed_filename.strip().lower() in self._renamed

    def is_url_done(self, url: str) -> bool:
        """
        Return True if this Source URL has already been processed.
        """
        if not url:
            return False
        return url.strip().lower() in self._urls

    def add_entry(self, original_filename: str, renamed_filename: str, source_url: str = "") -> None:
        """
        Record a completed original → renamed mapping.
        Updates in-memory sets AND persists to disk immediately.
        """
        orig_key    = original_filename.strip().lower()
        renamed_key = renamed_filename.strip().lower()
        url_key     = source_url.strip().lower()

        # Guard: don't write duplicates into the sheet itself (check renamed/url)
        # Note: multiple originals might map to same URL, but renamed is always unique.
        if renamed_key and renamed_key in self._renamed:
             log.debug(f"Excel tracker: renamed name already exists: {renamed_filename}")
             return

        if orig_key:
            self._originals.add(orig_key)
        if renamed_key:
            self._renamed.add(renamed_key)
        if url_key:
            self._urls.add(url_key)

        m = _TAMILPSD_RE.search(renamed_filename)
        if m:
            self.max_counter = max(self.max_counter, int(m.group(1)))

        self._save(original_filename.strip(), renamed_filename.strip(), source_url.strip())
        log.info(f"📊 Excel logged: {original_filename!r} → {renamed_filename!r} (URL: {source_url[:40]}...)")

    def stats(self) -> str:
        return (
            f"{len(self._originals)} originals tracked | "
            f"{len(self._renamed)} renamed tracked | "
            f"{len(self._urls)} URLs tracked | "
            f"max counter: tamilpsd-{self.max_counter:04d}"
        )
