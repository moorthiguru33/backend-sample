"""
cleanup_duplicates.py
══════════════════════════════════════════════════════════════════════
PURPOSE
  Move all EXTRA tamilpsd files from Google Drive (and local
  preview_image/) that are NOT recorded in rename_log.xlsx into a
  'duplicates/' subfolder.

  rename_log.xlsx is the SINGLE SOURCE OF TRUTH.
  Every tamilpsd file that is NOT in that log is a duplicate / stale
  upload and should be moved to the duplicates/ folder for review.

HOW TO RUN
  python scripts/cleanup_duplicates.py [--dry-run] [--drive] [--preview]

  --dry-run   Show what WOULD be moved without actually moving
  --drive     Clean Google Drive (requires GOOGLE_* env vars)
  --preview   Clean local preview_image/ folder
  (default: --dry-run --drive --preview)

WHAT IT DOES
  1. Reads rename_log.xlsx  → builds set of VALID tamilpsd names
  2. Scans Google Drive     → moves files NOT in valid set to duplicates/
  3. Scans preview_image/   → moves files NOT in valid set to duplicates/
  4. Prints a full report
══════════════════════════════════════════════════════════════════════
"""

import argparse
import logging
import os
import sys
import time as _time
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

_REPO_ROOT = Path(__file__).parent.parent
RENAME_LOG  = str(_REPO_ROOT / "rename_log.xlsx")
PREVIEW_DIR = _REPO_ROOT / "preview_image"


# ── Read rename_log ────────────────────────────────────────────────────────

def load_valid_names(rename_log_path: str) -> set:
    """
    Return a set of lowercase filenames (e.g. 'tamilpsd-0001.psd')
    that exist in rename_log.xlsx — these are VALID / wanted files.
    """
    wb = openpyxl.load_workbook(rename_log_path, read_only=True)
    ws = wb.active
    valid = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        renamed = row[1]
        if renamed:
            valid.add(str(renamed).strip().lower())
    wb.close()
    log.info(f"✅ rename_log.xlsx: {len(valid)} valid tamilpsd file names loaded")
    return valid


# ── Drive helpers ──────────────────────────────────────────────────────────

def _build_drive_service():
    """Build and return an authenticated Drive v3 service."""
    from config import GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET, GOOGLE_REFRESH_TOKEN

    import google.auth.transport.requests
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    SCOPES    = ["https://www.googleapis.com/auth/drive"]
    TOKEN_URI = "https://oauth2.googleapis.com/token"

    creds = Credentials(
        token=None,
        refresh_token=GOOGLE_REFRESH_TOKEN,
        token_uri=TOKEN_URI,
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=SCOPES,
    )
    creds.refresh(google.auth.transport.requests.Request())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _list_files(svc, folder_id):
    """Return all non-folder files directly inside folder_id."""
    items, page_token = [], None
    while True:
        params = dict(
            q=(
                f"'{folder_id}' in parents and trashed=false "
                "and mimeType!='application/vnd.google-apps.folder'"
            ),
            fields="nextPageToken,files(id,name,parents)",
            pageSize="1000",
        )
        if page_token:
            params["pageToken"] = page_token
        r = svc.files().list(**params).execute()
        items.extend(r.get("files", []))
        page_token = r.get("nextPageToken")
        if not page_token:
            break
    return items


def _list_subfolders(svc, folder_id):
    """Return all subfolder items directly inside folder_id."""
    items, page_token = [], None
    while True:
        params = dict(
            q=(
                f"'{folder_id}' in parents and trashed=false "
                "and mimeType='application/vnd.google-apps.folder'"
            ),
            fields="nextPageToken,files(id,name)",
            pageSize="1000",
        )
        if page_token:
            params["pageToken"] = page_token
        r = svc.files().list(**params).execute()
        items.extend(r.get("files", []))
        page_token = r.get("nextPageToken")
        if not page_token:
            break
    return items


def _find_or_create_duplicates_folder(svc, parent_folder_id: str) -> str:
    """Find or create a 'duplicates' subfolder inside parent_folder_id."""
    folder_name = "duplicates"
    q = (
        f"name='{folder_name}' and '{parent_folder_id}' in parents "
        f"and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    result = svc.files().list(q=q, fields="files(id,name)").execute()
    files = result.get("files", [])

    if files:
        log.info(f"📁 Found existing 'duplicates' folder: {files[0]['id']}")
        return files[0]["id"]

    # Create
    metadata = {
        "name":     folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents":  [parent_folder_id],
    }
    folder = svc.files().create(body=metadata, fields="id").execute()
    folder_id = folder["id"]
    log.info(f"📁 Created new 'duplicates' folder: {folder_id}")
    return folder_id


def _move_file_to_folder(svc, file_id: str, dest_folder_id: str) -> bool:
    """Move a file to a different folder on Drive."""
    try:
        # Get current parents
        file_info = svc.files().get(
            fileId=file_id, fields="parents"
        ).execute()
        current_parents = ",".join(file_info.get("parents", []))

        # Move
        svc.files().update(
            fileId=file_id,
            addParents=dest_folder_id,
            removeParents=current_parents,
            fields="id,parents",
        ).execute()
        return True
    except Exception as exc:
        log.warning(f"  Could not move file {file_id}: {exc}")
        return False


# ── Drive cleanup ──────────────────────────────────────────────────────────

def cleanup_drive(valid_names: set, dry_run: bool) -> None:
    """Scan all subfolders in GDRIVE_PSD_FOLDER and move extras to duplicates/."""
    from config import GDRIVE_PSD_FOLDER

    if not GDRIVE_PSD_FOLDER:
        log.error("Missing GDRIVE_PSD_FOLDER env var — cannot clean Drive.")
        return

    svc = _build_drive_service()

    log.info("🔍 Scanning Google Drive for ALL tamilpsd files…")

    total_scanned = 0
    total_moved   = 0
    total_kept    = 0
    to_move       = []

    # Root-level files
    for f in _list_files(svc, GDRIVE_PSD_FOLDER):
        total_scanned += 1
        if f["name"].lower() not in valid_names:
            to_move.append((f["id"], f["name"], "root"))
        else:
            total_kept += 1

    # Subfolders (skip 'duplicates' and 'final' folders)
    for sf in _list_subfolders(svc, GDRIVE_PSD_FOLDER):
        sf_name_lower = sf["name"].lower()
        if sf_name_lower in ("duplicates", "final"):
            log.info(f"  📁 Skipping '{sf['name']}' folder (reserved)")
            continue
        log.info(f"  📁 Scanning subfolder: {sf['name']}")
        for f in _list_files(svc, sf["id"]):
            total_scanned += 1
            if f["name"].lower() not in valid_names:
                to_move.append((f["id"], f["name"], sf["name"]))
            else:
                total_kept += 1

    log.info(f"\n{'='*60}")
    log.info(f"Drive scan complete:")
    log.info(f"  Total files scanned : {total_scanned}")
    log.info(f"  Valid (keep)        : {total_kept}")
    log.info(f"  Duplicates (move)   : {len(to_move)}")
    log.info(f"{'='*60}")

    if dry_run:
        log.info("DRY RUN — showing first 50 files that WOULD be moved to duplicates/:")
        for fid, name, folder in to_move[:50]:
            log.info(f"  [{folder}] {name}")
        if len(to_move) > 50:
            log.info(f"  ... and {len(to_move)-50} more")
        return

    if not to_move:
        log.info("✅ No duplicate files found — Drive is clean!")
        return

    # Create/find duplicates folder
    dup_folder_id = _find_or_create_duplicates_folder(svc, GDRIVE_PSD_FOLDER)
    if not dup_folder_id:
        log.error("❌ Could not create duplicates folder — aborting")
        return

    log.info(f"📦 Moving {len(to_move)} duplicate files to duplicates/ folder…")
    log.info(f"   (This may take several minutes for large batches)")

    for i, (fid, name, folder) in enumerate(to_move, 1):
        if _move_file_to_folder(svc, fid, dup_folder_id):
            total_moved += 1
        if i % 50 == 0:
            log.info(f"  ✅ Moved {i}/{len(to_move)} ({total_moved} success)…")
            _time.sleep(0.5)   # avoid Drive API rate limit

    log.info(f"✅ Drive cleanup done: {total_moved} files moved to duplicates/, {total_kept} files kept")


# ── Preview folder cleanup ─────────────────────────────────────────────────

def cleanup_preview(valid_names: set, dry_run: bool) -> None:
    """Move preview_image/*.webp files NOT in valid_names to a local duplicates subfolder."""
    if not PREVIEW_DIR.exists():
        log.info("preview_image/ folder not found — skipping")
        return

    webp_files = list(PREVIEW_DIR.glob("*.webp"))
    log.info(f"\n🔍 preview_image/: {len(webp_files)} .webp files found")

    to_move = []
    to_keep = []

    for f in webp_files:
        # Preview is named e.g. tamilpsd-0001.webp
        # valid_names contains tamilpsd-0001.psd / tamilpsd-0001.png etc.
        # Match by the numeric part: tamilpsd-XXXX
        stem = f.stem.lower()   # e.g. "tamilpsd-0001"
        # Check if ANY valid name starts with this stem
        if any(v.startswith(stem + ".") for v in valid_names):
            to_keep.append(f)
        else:
            to_move.append(f)

    log.info(f"  Valid (keep)           : {len(to_keep)}")
    log.info(f"  Duplicates (move)      : {len(to_move)}")

    if dry_run:
        log.info("DRY RUN — showing first 50 preview files that WOULD be moved:")
        for f in to_move[:50]:
            log.info(f"  {f.name}")
        if len(to_move) > 50:
            log.info(f"  ... and {len(to_move)-50} more")
        return

    if not to_move:
        log.info("✅ No duplicate preview files found — folder is clean!")
        return

    # Create local duplicates folder
    dup_dir = PREVIEW_DIR / "duplicates"
    dup_dir.mkdir(parents=True, exist_ok=True)

    log.info(f"📦 Moving {len(to_move)} duplicate preview files to preview_image/duplicates/…")
    moved = 0
    for f in to_move:
        try:
            dest = dup_dir / f.name
            f.rename(dest)
            moved += 1
        except Exception as exc:
            log.warning(f"  Could not move {f.name}: {exc}")

    log.info(f"✅ Preview cleanup done: {moved} moved to duplicates/, {len(to_keep)} kept")
    log.info("ℹ️  Remember to git commit & push after cleanup!")


# ── Report latest uploads ──────────────────────────────────────────────────

def show_latest_uploads(rename_log_path: str, n: int = 20) -> None:
    """Print the last N entries from rename_log (latest uploads)."""
    wb  = openpyxl.load_workbook(rename_log_path, read_only=True)
    ws  = wb.active
    rows = [(r[0], r[1]) for r in ws.iter_rows(min_row=2, values_only=True) if r[0] or r[1]]
    wb.close()

    log.info(f"\n{'='*60}")
    log.info(f"Latest {n} uploads in rename_log.xlsx:")
    log.info(f"{'Original (forpsd)':<35} → {'TamilPSD name'}")
    log.info(f"{'-'*60}")
    for orig, renamed in rows[-n:]:
        log.info(f"  {str(orig):<33} → {renamed}")
    log.info(f"{'='*60}")
    log.info(f"Total logged entries: {len(rows)}")


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Cleanup duplicate tamilpsd files (move to duplicates/ folder)")
    parser.add_argument("--dry-run",  action="store_true", default=True,
                        help="Show what would be moved (default: True)")
    parser.add_argument("--no-dry-run", dest="dry_run", action="store_false",
                        help="Actually move files to duplicates/")
    parser.add_argument("--drive",   action="store_true", default=True,
                        help="Clean Google Drive (default: True)")
    parser.add_argument("--preview", action="store_true", default=True,
                        help="Clean preview_image/ folder (default: True)")
    parser.add_argument("--show-latest", type=int, default=20,
                        help="Show latest N uploads from rename_log (default: 20)")
    args = parser.parse_args()

    if args.dry_run:
        log.info("🔵 DRY RUN MODE — nothing will be moved")
    else:
        log.info("🟡 LIVE MODE — duplicate files WILL be moved to duplicates/ folder!")
        log.info("   (Files are preserved for review, not deleted)")

    # Load valid names
    valid_names = load_valid_names(RENAME_LOG)

    # Show latest uploads
    show_latest_uploads(RENAME_LOG, args.show_latest)

    # Clean Drive
    if args.drive:
        cleanup_drive(valid_names, args.dry_run)

    # Clean preview_image/
    if args.preview:
        cleanup_preview(valid_names, args.dry_run)

    log.info("\n✅ Cleanup script finished.")
    if args.dry_run:
        log.info("Run with --no-dry-run to actually move the files to duplicates/.")


if __name__ == "__main__":
    main()
